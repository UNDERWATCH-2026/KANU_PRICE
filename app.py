import streamlit as st
import pandas as pd
import altair as alt
from supabase import create_client
from datetime import datetime, timedelta
import hashlib   # ✅ 반드시 추가

# =========================
# 0️⃣ 기본 설정
# =========================
st.set_page_config(page_title="Capsule Price Intelligence", layout="wide")

st.markdown("""
<style>
button[data-baseweb="tab"] {
    font-size: 15px;
    padding: 10px 30px;   /* 좌우 간격 조절 */
}
</style>
""", unsafe_allow_html=True)

def mk_widget_key(prefix: str, product_url: str, scope: str) -> str:
    raw = f"{prefix}|{product_url}|{scope}"
    return prefix + "_" + hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


# =========================
# 1️⃣ Supabase 설정
# =========================
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_ANON_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# =========================
# 2️⃣ 데이터 로딩
# =========================


def load_product_summary():
    cols = [
        "product_url",
        "brand",
        "category1",
        "category2",
        "product_name",
        "current_unit_price",
        "normal_unit_price",
        "is_discount",
        "first_seen_date",
        "last_seen_date",
        "event_count",
        "product_event_status",
        "is_new_product",
        "brew_type_kr",
        "capsule_count",        # ✅ 추가
    ]
    res = supabase.table("product_price_summary_enriched").select(", ".join(cols)).execute()
    df = pd.DataFrame(res.data)
    # 🔥 로딩 시점에 URL 정제
    df["product_url"] = df["product_url"].astype(str).str.strip("_").str.strip()
    return df

@st.cache_data(ttl=300)
def load_events(product_url: str):
    res = (
        supabase.table("product_all_events")
        .select("date, unit_price, event_type")
        .eq("product_url", product_url)
        .order("date", desc=True)
        .execute()
    )
    return pd.DataFrame(res.data)

@st.cache_data(ttl=300)
def load_lifecycle_events(product_url: str):
    res = (
        supabase.table("product_lifecycle_events")
        .select("date, lifecycle_event")
        .eq("product_url", product_url)
        .order("date", desc=True)
        .execute()
    )
    return pd.DataFrame(res.data)

# =========================
# 🚀 BULK 로딩 (API 최소화용)
# =========================

@st.cache_data(ttl=300)
def load_events_bulk(product_urls, date_from, date_to):
    empty = pd.DataFrame(columns=["product_url", "date", "unit_price", "event_type"])
    if not product_urls:
        return empty
    res = (
        supabase.table("product_all_events")
        .select("product_url, date, unit_price, event_type")
        .in_("product_url", product_urls)
        .gte("date", date_from.strftime("%Y-%m-%d"))
        .lte("date", date_to.strftime("%Y-%m-%d"))
        .execute()
    )
    if not res.data:
        return empty
    return pd.DataFrame(res.data)



@st.cache_data(ttl=300)
def load_lifecycle_bulk(product_urls, date_from, date_to):
    empty = pd.DataFrame(columns=["product_url", "date", "lifecycle_event"])
    if not product_urls:
        return empty
    res = (
        supabase.table("product_lifecycle_events")
        .select("product_url, date, lifecycle_event")
        .in_("product_url", product_urls)
        .gte("date", date_from.strftime("%Y-%m-%d"))
        .lte("date", date_to.strftime("%Y-%m-%d"))
        .execute()
    )
    if not res.data:
        return empty
    return pd.DataFrame(res.data)



@st.cache_data(ttl=300)
def load_raw_unit_bulk(product_urls, date_from, date_to):
    empty = pd.DataFrame(columns=["product_url", "date", "unit_normal_price"])
    if not product_urls:
        return empty
    res = (
        supabase.table("raw_daily_prices_unit")
        .select("product_url, date, unit_normal_price")
        .in_("product_url", product_urls)
        .gte("date", date_from.strftime("%Y-%m-%d"))
        .lte("date", date_to.strftime("%Y-%m-%d"))
        .execute()
    )
    if not res.data:
        return empty
    return pd.DataFrame(res.data)
# =========================
# 2-1️⃣ 질문 로그 저장
# =========================
def save_question_log(question: str, q_type: str, used_llm: bool, answer: str = None, filters: dict = None):
    try:
        log_data = {
            "question_text": question,
            "question_type": q_type,
            "used_llm": used_llm,
            "created_at": datetime.now().isoformat()
        }
        supabase.table("question_logs").insert(log_data).execute()
    except Exception as e:
        print("로그 저장 실패:", e)


# =========================
# 2-2️⃣ 질문 처리 함수들
# =========================

def normalize_brand_name(brand_query: str) -> str:
    brand_query = brand_query.lower().strip()
    brand_mapping = {
        "카누": "카누 바리스타",
        "카누바리스타": "카누 바리스타",
        "카누 바리스타": "카누 바리스타",
        "카누돌체구스토": "카누 돌체구스토",
        "카누 돌체구스토": "카누 돌체구스토",
    
        "네스프레소": "네스프레소",
        "네슬레": "네슬레",
    
        # 🔥 추가
        "일리": "일리카페",
        "일리카페": "일리카페",
    
        "돌체구스토": "돌체구스토",
        "돌체": "돌체구스토",
    
        "스타벅스": "스타벅스",
    }
    for key, value in brand_mapping.items():
        if key.replace(" ", "") == brand_query.replace(" ", ""):
            return value
    return brand_query

def extract_brand_from_question(q: str, df_all: pd.DataFrame) -> list:
    q_lower = q.lower()
    brands = df_all["brand"].dropna().unique().tolist()
    matched_brands = []

    normalized = None
    for key in ["카누","네스프레소","네슬레","일리","일리카페","돌체","돌체구스토","스타벅스"]:
        if key in q_lower:
            normalized = normalize_brand_name(key)
            break

    for brand in brands:
        brand_lower = brand.lower()

        if brand_lower in q_lower:
            matched_brands.append(brand)

        elif normalized and brand_lower == normalized.lower():
            matched_brands.append(brand)

    return matched_brands if matched_brands else None
    
    for brand in brands:
        normalized = normalize_brand_name(q_lower)
        if brand.lower() == normalized.lower():
            matched_brands.append(brand)
    if matched_brands:
        return matched_brands
    for brand in brands:
        brand_lower = brand.lower()
        brand_parts = brand_lower.split()
        for part in brand_parts:
            if len(part) >= 2 and part in q_lower:
                if brand not in matched_brands:
                    matched_brands.append(brand)
                break
    return matched_brands if matched_brands else None

def extract_product_name_from_question(q: str) -> list:
    exclude_words = [
        "할인", "기간", "언제", "얼마", "가격", "제품",
        "최저가", "최고가", "신제품", "품절", "복원", "중", "는", "은", "의",
        "신상", "출시", "새로", "신규", "새로운", "처음",
        "할인가", "정상가", "변동", "상승", "증가", "하락",
        "비싼", "싼", "저렴", "최근", "알려줘", "보여줘",
        "있어", "없어", "언제부터", "언제까지", "기간은", "얼마야",
        "날짜", "할인률", "할인율", "높은", "낮은", "순위", "순서",
        "상위", "하위", "개", "위", "등", "번째", "최대", "최소",
        "많은", "적은", "전체", "모든", "모두", "모아", "알려",
    ]
    import re as _re
    words = q.split()
    product_keywords = []
    for word in words:
        # 숫자만 있거나 숫자+단위(개/위/등)면 제외
        if _re.fullmatch(r"[0-9]+[개위등번]*", word):
            continue
        if len(word) >= 2 and not any(ex in word for ex in exclude_words):
            product_keywords.append(word)
    return product_keywords

def extract_top_n(q: str):
    """질문에서 상위/하위 N 추출.
    반환: (n, direction, mode)
      direction: "top" | "bottom"
      mode: "count" (N개) | "rank" (N위 - 동점 모두 포함)
    """
    import re
    # "상위/하위 N위", "N위"
    m = re.search(r'(상위|하위)?\s*(\d+)\s*(위|등)', q)
    if m:
        n = int(m.group(2))
        direction = "bottom" if m.group(1) == "하위" else "top"
        return n, direction, "rank"
    # "상위/하위 N개", "N개"
    m = re.search(r'(상위|하위|최대|최고|최저|top)?\s*(\d+)\s*개', q)
    if m:
        n = int(m.group(2))
        direction = "bottom" if m.group(1) == "하위" else "top"
        return n, direction, "count"
    return None, None, None

def classify_intent(q: str):
    q = q.lower()

    if "날짜" in q:
        if "품절" in q and "복원" in q:
            return "OUT_AND_RESTORE_DATES"
        if "품절" in q:
            return "OUT_DATES"
        if "복원" in q:
            return "RESTORE_DATES"
        if "신제품" in q or "출시" in q:
            return "NEW_DATES"

    if any(word in q for word in ["신제품", "새로", "신규", "출시", "신상"]) and "품절" in q:
        return "NEW_AND_OUT"

    if "품절" in q and ("복원" in q or "재입고" in q):
        if any(word in q for word in ["후", "다시", "그후"]):
            return "RESTORE"
        return "OUT_AND_RESTORE"

    if "할인" in q and ("기간" in q or "언제" in q):
        return "DISCOUNT_PERIOD"
    if "할인" in q and any(w in q for w in ["률", "율", "퍼센트", "%", "높은", "최대", "가장 많이"]):
        return "DISCOUNT_RATE"
    if "할인" in q or "행사" in q:
        return "DISCOUNT"
    if any(word in q for word in ["신제품", "새롭게", "새로", "신규", "출시", "새로운", "처음", "신상"]):
        return "NEW"
    if any(word in q for word in ["가장 싼", "제일 싼", "제일 저렴한", "가장 저렴한", "최저가"]):
        return "PRICE_MIN"
    if any(word in q for word in ["가장 비싼", "제일 비싼", "최고가"]):
        return "PRICE_MAX"
    if "정상가" in q and ("변동" in q or "상승" in q or "인상" in q or "올랐" in q or "인하" in q or "내렸" in q):
        return "NORMAL_CHANGE"
    if any(word in q for word in ["상승", "증가"]) and "않" not in q:
        return "PRICE_UP"
    if "변동" in q or "많이 바뀐" in q:
        return "VOLATILITY"
    if any(word in q for word in ["복원", "재입고", "입고", "돌아온"]):
        return "RESTORE"
    if "품절" in q:
        return "OUT"


    return "UNKNOWN"

def extract_period_from_question(q: str, base_date=None):
    import re
    today = base_date if base_date else datetime.today()
    q_lower = q.lower()

    if any(w in q_lower for w in ["최근 7일", "최근 일주일", "최근 1주일"]):
        return today - timedelta(days=7), today, "최근 7일 내"
    if any(w in q_lower for w in ["최근 한 달", "최근 30일"]):
        return today - timedelta(days=30), today, "최근 1개월 내"
    if any(w in q_lower for w in ["최근 1년"]):
        return today - timedelta(days=365), today, "최근 1년 내"

    # ✅ "최근 N개월" 통합 처리
    m = re.search(r"최근\s*(\d+)\s*개월", q_lower)
    if m:
        n = int(m.group(1))
        label = "최근 1년 내" if n == 12 else f"최근 {n}개월 내"
        return today - timedelta(days=n * 30), today, label

    month_match = re.search(r"(\d{4})년\s*(\d{1,2})월", q)
    if month_match:
        year = int(month_match.group(1))
        month = int(month_match.group(2))
        from_dt = datetime(year, month, 1)
        to_dt = (datetime(year, month + 1, 1) - timedelta(days=1)) if month < 12 else datetime(year, 12, 31)
        return from_dt, to_dt, f"{year}년 {month}월"

    year_match = re.search(r"(\d{4})년", q)
    if year_match:
        year = int(year_match.group(1))
        return datetime(year, 1, 1), datetime(year, 12, 31), f"{year}년"

    return None

def extract_brew_type(q: str, df_all: pd.DataFrame):
    q_lower = q.lower()
    brew_list = df_all["brew_type_kr"].dropna().unique().tolist()
    # 브랜드명은 brew_type으로 잘못 매칭되지 않도록 제외
    brand_list = [str(b).lower() for b in df_all["brand"].dropna().unique().tolist()]
    # 1단계: brew가 질문에 포함 (예: "에스프레소 질문" → "에스프레소")
    for brew in brew_list:
        if brew and brew.lower() in q_lower:
            # brew 값이 브랜드명과 동일하면 스킵
            if brew.lower() in brand_list:
                continue
            return brew
    # 2단계: 질문이 brew에 포함 (예: "에스프레소" → "에스프레소 (LTO)")
    for brew in brew_list:
        if brew and q_lower in brew.lower():
            if brew.lower() in brand_list or q_lower in brand_list:
                continue
            return brew
    return None

def _apply_top_n(products, product_details, top_n):
    """top_n = (n, "top"/"bottom") 에 따라 products/product_details 슬라이싱"""
    if not top_n:
        return products, product_details
    n, direction = top_n
    if direction == "bottom":
        sliced = products[-n:]
    else:
        sliced = products[:n]
    sliced_set = set(sliced)
    return sliced, {k: v for k, v in product_details.items() if k in sliced_set}

def execute_rule(intent, question, df_summary, date_from=None, date_to=None, top_n=None):
    result = _execute_rule_inner(intent, question, df_summary, date_from, date_to)
    if top_n and isinstance(result, dict) and result.get("type") == "product_list":
        import re as _re
        n, direction, mode = top_n  # mode: "count" | "rank"
        products = result.get("products", [])
        product_details = result.get("product_details", {})

        # df_summary 기준 메타 정보 (동점 시 가나다순)
        url_meta = {}
        for _, row in df_summary.iterrows():
            url_key = str(row["product_url"]).strip().lower()
            url_meta[url_key] = (
                str(row.get("brand") or ""),
                str(row.get("category1") or ""),
                str(row.get("category2") or ""),
                str(row.get("product_name") or ""),
            )

        # products는 _inner에서 이미 수치 정렬된 순서 → 원래 순서(idx) 유지
        # 동점(같은 detail 값)끼리만 가나다순 보조 정렬
        detail_of = lambda url: product_details.get(url, "")

        # 동점 그룹: detail 동일 → 같은 rank, 가나다순 정렬
        # 다른 detail → 원래 products 순서(idx) 유지
        detail_vals = [detail_of(u) for u in products]
        unique_details_ordered = []
        for d in detail_vals:
            if d not in unique_details_ordered:
                unique_details_ordered.append(d)
        rank_map = {d: i for i, d in enumerate(unique_details_ordered)}

        def sort_key(url):
            rank = rank_map[detail_of(url)]
            meta = url_meta.get(url, ("", "", "", ""))
            return (rank, meta[0], meta[1], meta[2], meta[3])

        sorted_products = sorted(products, key=sort_key)

        if mode == "rank":
            # N위까지 - N위와 동일한 순위 제품 모두 포함
            sorted_ranks = sorted(set(rank_map[detail_of(u)] for u in products))
            cutoff_rank = sorted_ranks[min(n - 1, len(sorted_ranks) - 1)] if direction != "bottom" else sorted_ranks[-(n)]
            if direction == "bottom":
                sliced = [u for u in sorted_products if rank_map[detail_of(u)] >= cutoff_rank]
            else:
                sliced = [u for u in sorted_products if rank_map[detail_of(u)] <= cutoff_rank]
        else:
            # N개 - 단순 슬라이싱
            if direction == "bottom":
                sliced = sorted_products[-n:]
            else:
                sliced = sorted_products[:n]

        sliced_set = set(sliced)
        result["products"] = sliced
        result["product_details"] = {k: v for k, v in product_details.items() if k in sliced_set}
        result["text"] = _re.sub(r'[(][0-9]+개', f'({len(sliced)}개', result["text"])
    return result

def _execute_rule_inner(intent, question, df_summary, date_from=None, date_to=None):

    df_work = df_summary.copy()
    brands = extract_brand_from_question(question, df_summary)

    if brands:
        df_work = df_work[
            df_work["brand"].astype(str).str.strip().isin(brands)
        ]
    # 🔧 추가
    all_keywords = extract_product_name_from_question(question)

    question_period = extract_period_from_question(question)

    if question_period:
        q_from, q_to, period_label = question_period
        date_from = q_from
        date_to = q_to
    else:
        if date_from and date_to:
            period_label = f"{date_from.strftime('%Y-%m-%d')} ~ {date_to.strftime('%Y-%m-%d')}"
        else:
            period_label = "조회 기간 내"

    brew_condition = extract_brew_type(question, df_summary)

    if brew_condition:
        filtered = df_work[
            df_work["brew_type_kr"].str.contains(brew_condition, na=False)
        ]
        if not filtered.empty:
            df_work = filtered

        if all_keywords and df_work.empty:
            keywords_str = ", ".join(all_keywords)
            return f"'{keywords_str}'에 해당하는 제품이 없습니다."

    # =========================
    # 🔥 할인 기간 조회
    # =========================
    if intent == "DISCOUNT_PERIOD":
        # ✅ 수정1: 제품명 키워드 필터 추가
        product_keywords = extract_product_name_from_question(question)
        brand_words = [b.lower().replace(" ", "") for b in (brands or [])]
        product_keywords = [
            kw for kw in product_keywords
            if kw.lower().replace(" ", "") not in brand_words
            and kw not in ["할인", "기간", "언제", "알려줘", "보여줘"]
        ]
        
        if product_keywords:
            for kw in product_keywords:
                mask = _norm_series(df_work["product_name"]).str.contains(kw, case=False)
                df_work = df_work[mask]
                if df_work.empty:
                    break
    
        results = []
        for _, row in df_work.iterrows():
            res = supabase.rpc(
                "get_discount_periods_in_range",
                {
                    "p_product_url": row["product_url"],
                    "p_date_from": (date_from if date_from else datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d"),
                    "p_date_to": (date_to if date_to else datetime.now()).strftime("%Y-%m-%d"),
                }
            ).execute()
            discount_periods = res.data if res.data else []
            if discount_periods:
                for period in discount_periods:
                    # 할인가 조회
                    discount_price_res = (
                        supabase.table("product_all_events")
                        .select("unit_price")
                        .eq("product_url", row["product_url"])
                        .eq("event_type", "DISCOUNT")
                        .gte("date", period["discount_start_date"])
                        .lte("date", period["discount_end_date"])
                        .limit(1)
                        .execute()
                    )
                    discount_price = float(discount_price_res.data[0]["unit_price"]) if discount_price_res.data else None
    
                    # ✅ 수정2: 정상가 조회 - 할인 직전 NORMAL 우선, 없으면 summary fallback
                    normal_price = None
                    normal_price_res = (
                        supabase.table("product_all_events")
                        .select("unit_price")
                        .eq("product_url", row["product_url"])
                        .eq("event_type", "NORMAL")
                        .lt("date", period["discount_start_date"])
                        .order("date", desc=True)
                        .limit(1)
                        .execute()
                    )
                    if normal_price_res.data:
                        normal_price = float(normal_price_res.data[0]["unit_price"])
                    
                    # fallback: summary의 normal_unit_price
                    if not normal_price or normal_price <= 0:
                        v = float(row.get("normal_unit_price") or 0)
                        if v > 0:
                            normal_price = v
    
                    # ✅ 수정3: 가격 표시 형식 통일
                    if normal_price and discount_price and normal_price > discount_price:
                        discount_rate = (normal_price - discount_price) / normal_price * 100
                        price_info = f"💰 정상가: {normal_price:,.1f}원 → 할인가: {discount_price:,.1f}원 ({discount_rate:.0f}% 할인)"
                    elif discount_price:
                        price_info = f"💰 할인가: {discount_price:,.1f}원 (정상가 정보 없음)"
                    else:
                        price_info = ""
    
                    url = str(row["product_url"]).strip().lower()
                    disp_start = max(period['discount_start_date'], date_from.strftime("%Y-%m-%d") if date_from else period['discount_start_date'])
                    disp_end = min(period['discount_end_date'], date_to.strftime("%Y-%m-%d") if date_to else period['discount_end_date'])
                    
                    # ✅ 수정4: detail_str 형식도 통일
                    if normal_price and discount_price and normal_price > discount_price:
                        discount_rate = (normal_price - discount_price) / normal_price * 100
                        detail_str = f"📅 {disp_start} ~ {disp_end} | 💰 {normal_price:,.1f}원 → {discount_price:,.1f}원 ({discount_rate:.0f}% 할인)"
                    elif discount_price:
                        detail_str = f"📅 {disp_start} ~ {disp_end} | 💰 할인가: {discount_price:,.1f}원"
                    else:
                        detail_str = f"📅 {disp_start} ~ {disp_end}"
    
                    results.append({
                        "text": f"• {row['brand']} - {row['product_name']}\n"
                                f"  📅 할인 기간: {disp_start} ~ {disp_end}\n"
                                f"{price_info}",
                        "product_url": url,
                        "detail": detail_str,
                    })
    
        if not results:
            return "해당 제품의 할인 기간 정보가 없습니다."
        
        product_details = {}
        for r in results:
            url = r["product_url"]
            if url in product_details:
                product_details[url] += " → " + r["detail"]
            else:
                product_details[url] = r["detail"]
        
        seen = set()
        unique_products = []
        for r in results:
            if r["product_url"] not in seen:
                seen.add(r["product_url"])
                unique_products.append(r["product_url"])
        
        return {
            "type": "product_list",
            "text": f"할인 기간 정보 ({len(unique_products)}개)",
            "products": unique_products,
            "product_details": product_details,
        }
    
    if intent == "DISCOUNT_RATE":
        # 조회 기간 내 할인율 최대 제품 조회
        res_d = supabase.table("product_all_events").select("product_url, unit_price, date").eq("event_type", "DISCOUNT")
        res_n = supabase.table("product_all_events").select("product_url, unit_price, date").eq("event_type", "NORMAL")
        if date_from:
            res_d = res_d.gte("date", date_from.strftime("%Y-%m-%d"))
            res_n = res_n.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_d = res_d.lte("date", date_to.strftime("%Y-%m-%d"))
            res_n = res_n.lte("date", date_to.strftime("%Y-%m-%d"))
        res_d = res_d.execute()
        res_n = res_n.execute()

        df_discount = pd.DataFrame(res_d.data) if res_d.data else pd.DataFrame()
        df_normal = pd.DataFrame(res_n.data) if res_n.data else pd.DataFrame()
        
        if not df_discount.empty:
            df_discount["date"] = pd.to_datetime(df_discount["date"])
        
            # 조회기간 필터
            df_discount = df_discount[
                (df_discount["date"] >= pd.Timestamp(date_from)) &
                (df_discount["date"] <= pd.Timestamp(date_to))
            ]
        
        if not df_normal.empty:
            df_normal["date"] = pd.to_datetime(df_normal["date"])
        
            # 조회기간 필터
            df_normal = df_normal[
                (df_normal["date"] >= pd.Timestamp(date_from)) &
                (df_normal["date"] <= pd.Timestamp(date_to))
            ]
        
        # 🔥 이 위치로 이동
        if df_discount.empty:
            return "해당 기간 내 할인 이벤트가 없습니다."
        
        # 제품별 최저 할인가 + 날짜
        discount_map = {}
        discount_date_map = {}
        
        for _, r in df_discount.iterrows():
            key = str(r["product_url"]).strip().lower()
            p = float(r["unit_price"]) if r["unit_price"] else 0
        
            if p > 0:
                if key not in discount_map or p < discount_map[key]:
                    discount_map[key] = p
                    discount_date_map[key] = r["date"]

        # 제품별 정상가 (할인 직전 or 기간 내 NORMAL)
        normal_map = {}
        for _, r in df_normal.iterrows():
            key = str(r["product_url"]).strip().lower()
            p = float(r["unit_price"]) if r["unit_price"] else 0
            if p > 0:
                if key not in normal_map or p > normal_map[key]:
                    normal_map[key] = p

        # df_work 기준 URL만 사용 (브랜드/키워드 필터 반영)
        df_work_urls = set(df_work["product_url"].str.strip().str.lower().tolist())

        # summary의 normal_unit_price를 fallback 정상가로 미리 수집
        summary_normal = {}
        for _, row in df_work.iterrows():
            url_key = str(row["product_url"]).strip().lower()
            v = float(row.get("normal_unit_price") or 0)
            if v > 0:
                summary_normal[url_key] = v

        # 기간 내 NORMAL로도 정상가 못 찾은 경우 기간 밖 최근 NORMAL 조회 (bulk)
        missing_normal_urls = [
            url for url in discount_map
            if url in df_work_urls and url not in normal_map and url not in summary_normal
        ]
        if missing_normal_urls:
            # 원본 URL로 변환
            url_orig_map = {str(r["product_url"]).strip().lower(): r["product_url"] for r in (res_d.data or [])}
            for url in missing_normal_urls:
                orig_url = url_orig_map.get(url, url)
                nr = supabase.table("product_all_events").select("unit_price").eq("product_url", orig_url).eq("event_type", "NORMAL").order("date", desc=True).limit(1).execute()
                if nr.data:
                    p = float(nr.data[0]["unit_price"]) if nr.data[0]["unit_price"] else 0
                    if p > 0:
                        summary_normal[url] = p

        # 할인율 계산
        rate_list = []
        for url, disc_price in discount_map.items():
            if url not in df_work_urls:
                continue
            norm_price = normal_map.get(url) or summary_normal.get(url)
            if norm_price and norm_price > disc_price:
                rate = (norm_price - disc_price) / norm_price * 100
                rate_list.append((url, disc_price, norm_price, rate))
            elif disc_price > 0:
                # 정상가 불명이면 할인율 0%로 포함 (제품은 보여주되 할인율 미상)
                rate_list.append((url, disc_price, None, 0.0))

        if not rate_list:
            return "할인율 계산 가능한 제품이 없습니다."

        rate_list.sort(key=lambda x: -x[3])

        urls = [r[0] for r in rate_list]
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(urls)].drop_duplicates(subset=["product_url"])

        results = []
        product_details = {}
        # 원본 URL 맵 (기간 조회용)
        url_orig_map_for_period = {str(r["product_url"]).strip().lower(): r["product_url"] for r in (res_d.data or [])}
        date_from_str2 = date_from.strftime("%Y-%m-%d") if date_from else (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
        date_to_str2 = date_to.strftime("%Y-%m-%d") if date_to else datetime.now().strftime("%Y-%m-%d")

        for url, disc_price, norm_price, rate in rate_list:
            results.append({"product_url": url})
            # 할인 기간 조회
            orig_url = url_orig_map_for_period.get(url, url)
            period_res2 = supabase.rpc("get_discount_periods_in_range", {
                "p_product_url": orig_url,
                "p_date_from": date_from_str2,
                "p_date_to": date_to_str2,
            }).execute()
            if period_res2.data:
                periods_str2 = "  /  ".join([
                    f"📅 {max(p['discount_start_date'], date_from_str2)} ~ {min(p['discount_end_date'], date_to_str2)}"
                    for p in period_res2.data
                ])
                period_detail = f"  |  {periods_str2}"
            else:
                disc_date = discount_date_map.get(url, "")
                period_detail = f"  |  📅 {disc_date}" if disc_date else ""
            if norm_price:
                product_details[url] = f"💰 {norm_price:,.1f}원 → {disc_price:,.1f}원 ({rate:.1f}%){period_detail}"
            else:
                product_details[url] = f"💰 {disc_price:,.1f}원 (정상가 미상){period_detail}"

        return {
            "type": "product_list",
            "text": f"{period_label} 할인율 높은 제품 ({len(results)}개, 높은 순)",
            "products": urls,
            "product_details": product_details,
        }

    if intent == "DISCOUNT" and not start_date:
        # df_work의 URL 목록으로 직접 조회 (브랜드/키워드 필터 완전 반영)
        df_work_dedup = df_work.drop_duplicates(subset=["product_url"])
        date_from_str = date_from.strftime("%Y-%m-%d") if date_from else (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
        date_to_str = date_to.strftime("%Y-%m-%d") if date_to else datetime.now().strftime("%Y-%m-%d")

        discount_map = {}  # orig_url -> [prices]
        for _, row in df_work_dedup.iterrows():
            orig_url = row["product_url"]
            res_d = (
                supabase.table("product_all_events")
                .select("unit_price")
                .eq("product_url", orig_url)
                .eq("event_type", "DISCOUNT")
                .gte("date", date_from_str)
                .lte("date", date_to_str)
                .execute()
            )
            if res_d.data:
                prices = [float(r["unit_price"]) for r in res_d.data if r["unit_price"] and float(r["unit_price"]) > 0]
                if prices:
                    discount_map[str(orig_url).strip().lower()] = prices

        df = df_work_dedup[df_work_dedup["product_url"].str.strip().str.lower().isin(discount_map.keys())]

        if df.empty:
            return None

        # res_discount placeholder (하위 호환용)
        class _FakeRes:
            data = list(discount_map.keys())
        res_discount = _FakeRes()

        df = df.drop_duplicates(subset=["product_url"])
        results = []
        product_details = {}
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            if url in discount_map:
                disc_price = min(discount_map[url])
                norm_price = float(row.get("normal_unit_price") or 0)
                if norm_price <= 0:
                    norm_price = float(row.get("current_unit_price") or 0)
                rate_str = f" ({(norm_price - disc_price) / norm_price * 100:.1f}%)" if norm_price > disc_price else ""
                norm_str = f"{norm_price:,.1f}원 → " if norm_price > 0 else ""
                # 할인 기간 조회
                period_res = supabase.rpc(
                    "get_discount_periods_in_range",
                    {
                        "p_product_url": row["product_url"],
                        "p_date_from": (date_from if date_from else datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d"),
                        "p_date_to": (date_to if date_to else datetime.now()).strftime("%Y-%m-%d"),
                    }
                ).execute()
                if period_res.data:
                    # 조회 기간으로 클리핑
                    date_from_str = date_from.strftime("%Y-%m-%d") if date_from else None
                    date_to_str = date_to.strftime("%Y-%m-%d") if date_to else None
                    periods_str = "  /  ".join([
                        f"📅 {max(p['discount_start_date'], date_from_str) if date_from_str else p['discount_start_date']} ~ {min(p['discount_end_date'], date_to_str) if date_to_str else p['discount_end_date']}"
                        for p in period_res.data
                    ])
                    detail = f"💰 {norm_str}{disc_price:,.1f}원{rate_str}  |  {periods_str}"
                else:
                    detail = f"💰 {norm_str}{disc_price:,.1f}원{rate_str}"
            else:
                disc_price = float(row["current_unit_price"])
                norm_price = float(row.get("normal_unit_price") or 0)
                rate_str = f" ({(norm_price - disc_price) / norm_price * 100:.1f}%)" if norm_price > disc_price else ""
                norm_str = f"{norm_price:,.1f}원 → " if norm_price > 0 else ""
                detail = f"💰 {norm_str}{disc_price:,.1f}원{rate_str}"
            results.append({"product_url": url})
            product_details[url] = detail

        if not results:
            return None
        return {
            "type": "product_list",
            "text": f"{period_label} 할인 제품 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }

    # =========================
    # 🔥 PRICE_MIN
    # =========================
    if intent == "PRICE_MIN":
    
        res = (
            supabase.table("product_all_events")
            .select("product_url, date, unit_price")
            .in_("product_url", df_work["product_url"].tolist())
            .gte("date", date_from.strftime("%Y-%m-%d"))
            .lte("date", date_to.strftime("%Y-%m-%d"))
            .execute()
        )
    
        if not res.data:
            return "가격 데이터가 없습니다."
    
        df_hist = pd.DataFrame(res.data)

        
        # 🔥 브랜드 필터 유지
        df_hist = df_hist[
            df_hist["product_url"].astype(str).str.strip().str.lower()
            .isin(df_work["product_url"].astype(str).str.strip().str.lower())
        ]
        

        df_hist["date"] = pd.to_datetime(df_hist["date"])
        df_hist["unit_price"] = df_hist["unit_price"].astype(float)
    
        # 조회기간 필터
        df_hist = df_hist[
            (df_hist["date"] >= pd.Timestamp(date_from)) &
            (df_hist["date"] <= pd.Timestamp(date_to))
        ]
    
        df_hist = df_hist[df_hist["unit_price"] > 0]
    
        if df_hist.empty:
            return "조회 기간 내 가격 데이터가 없습니다."
    
        # 🔥 기간 내 최저가 계산
        min_price = df_hist["unit_price"].min()
    
        df_low = df_hist[df_hist["unit_price"] == min_price]

        # 🔥 브랜드 강제 유지
        df_low = df_low[
            df_low["product_url"].astype(str).str.strip().str.lower()
            .isin(df_work["product_url"].astype(str).str.strip().str.lower())
        ]
            
        results = []
        product_details = {}
    
        for url in df_low["product_url"].unique():
    
            row = df_work[df_work["product_url"] == url]
            if row.empty:
                continue
    
            row = row.iloc[0]
    
            df_prod = df_low[df_low["product_url"] == url]
    
            sd = df_prod["date"].min().date()
            ed = df_prod["date"].max().date()
    
            categories = []
            if pd.notna(row.get("category1")) and row["category1"]:
                categories.append(row["category1"])
            if pd.notna(row.get("category2")) and row["category2"]:
                categories.append(row["category2"])
    
            category_str = f" [{' > '.join(categories)}]" if categories else ""
    
            url_key = str(url).strip().lower()
    
            results.append({"product_url": url_key})
    
            product_details[url_key] = (
                f"💰 최저가: {min_price:,.1f}원 ({sd} ~ {ed})"
            )
    
        return {
            "type": "product_list",
            "text": f"{period_label} 최저가 제품 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }
    
    
    # =========================
    # 🔥 PRICE_MAX
    # =========================
    if intent == "PRICE_MAX":
    
        res = (
            supabase.table("product_all_events")
            .select("product_url, date, unit_price")
            .in_("product_url", df_work["product_url"].tolist())
            .gte("date", date_from.strftime("%Y-%m-%d"))
            .lte("date", date_to.strftime("%Y-%m-%d"))
            .execute()
        )
    
        if not res.data:
            return "가격 데이터가 없습니다."
    
        df_hist = pd.DataFrame(res.data)
        # 🔥 브랜드 필터 유지
        df_hist = df_hist[
            df_hist["product_url"].astype(str).str.strip().str.lower()
            .isin(df_work["product_url"].astype(str).str.strip().str.lower())
        ]
        

        df_hist["date"] = pd.to_datetime(df_hist["date"])
        df_hist["unit_price"] = df_hist["unit_price"].astype(float)
    
        # 조회기간 필터
        df_hist = df_hist[
            (df_hist["date"] >= pd.Timestamp(date_from)) &
            (df_hist["date"] <= pd.Timestamp(date_to))
        ]
    
        df_hist = df_hist[df_hist["unit_price"] > 0]
    
        if df_hist.empty:
            return "조회 기간 내 가격 데이터가 없습니다."
    
        # 🔥 기간 내 최고가
        max_price = df_hist["unit_price"].max()
    
        df_hi = df_hist[df_hist["unit_price"] == max_price]
    
        results = []
        product_details = {}
    
        for url in df_hi["product_url"].unique():
    
            row = df_work[df_work["product_url"] == url]
            if row.empty:
                continue
    
            row = row.iloc[0]
    
            df_prod = df_hi[df_hi["product_url"] == url]
    
            sd = df_prod["date"].min().date()
            ed = df_prod["date"].max().date()
    
            url_key = str(url).strip().lower()
    
            results.append({"product_url": url_key})
    
            product_details[url_key] = (
                f"💰 최고가: {max_price:,.1f}원 ({sd} ~ {ed})"
            )
    
        return {
            "type": "product_list",
            "text": f"{period_label} 최고가 제품 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }

    # =========================
    # 🔥 날짜 전용: 품절 날짜
    # =========================
    if intent == "OUT_DATES":
        res = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "OUT_OF_STOCK")
        )
        if date_from:
            res = res.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res = res.lte("date", date_to.strftime("%Y-%m-%d"))
        res = res.execute()
        if not res.data:
            return "해당 기간 내 품절 이력이 없습니다."
        out_map = {}
        for r in res.data:
            key = str(r["product_url"]).strip().lower()
            out_map.setdefault(key, []).append(r["date"])
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(out_map.keys())]
        if df.empty:
            return "해당 제품의 품절 이력이 없습니다."
        results = []
        product_details = {}  # 🔥 추가
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            dates = sorted(out_map.get(url, []))
            date_str = ", ".join(dates)
            results.append({
                "text": f"• {row['brand']} - {row['product_name']}\n  ❌ 품절 날짜: {date_str}",
                "product_url": url
            })
            product_details[url] = f"❌ 품절 날짜: {date_str}"
        return {
            "type": "product_list",
            "text": f"품절 날짜 정보 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }

    # =========================
    # 🔥 날짜 전용: 복원 날짜
    # =========================
    if intent == "RESTORE_DATES":
        res = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "RESTOCK")
        )
        if date_from:
            res = res.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res = res.lte("date", date_to.strftime("%Y-%m-%d"))
        res = res.execute()
        if not res.data:
            return "해당 기간 내 복원 이력이 없습니다."
        restore_map = {}
        for r in res.data:
            key = str(r["product_url"]).strip().lower()
            restore_map.setdefault(key, []).append(r["date"])
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(restore_map.keys())]
        if df.empty:
            return "해당 제품의 복원 이력이 없습니다."
        results = []
        product_details = {}  # 🔥 추가
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            dates = sorted(restore_map.get(url, []))
            date_str = ", ".join(dates)
            results.append({
                "text": f"• {row['brand']} - {row['product_name']}\n  🔄 복원 날짜: {date_str}",
                "product_url": url
            })
            product_details[url] = f"🔄 복원 날짜: {date_str}"
        return {
            "type": "product_list",
            "text": f"복원 날짜 정보 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }

    # =========================
    # 🔥 날짜 전용: 신제품 날짜
    # =========================
    if intent == "NEW_DATES":
        res = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "NEW_PRODUCT")
        )
        if date_from:
            res = res.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res = res.lte("date", date_to.strftime("%Y-%m-%d"))
        res = res.execute()
        if not res.data:
            return "해당 기간 내 출시 이력이 없습니다."
        new_map = {}
        for r in res.data:
            key = str(r["product_url"]).strip().lower()
            new_map.setdefault(key, []).append(r["date"])
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(new_map.keys())]
        if df.empty:
            return "해당 제품의 출시 이력이 없습니다."
        results = []
        product_details = {}  # 🔥 추가
        for _, row in df.iterrows():
            dates = sorted(new_map[row["product_url"]])
            date_str = ", ".join(dates)
            url = str(row["product_url"])
            results.append({
                "text": f"• {row['brand']} - {row['product_name']}\n  🆕 출시 날짜: {date_str}",
                "product_url": url
            })
            product_details[url] = f"🆕 출시 날짜: {date_str}"  # 🔥 추가
        return {
            "type": "product_list",
            "text": f"출시 날짜 정보 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,  # 🔥 추가
        }

    # =========================
    # 🔥 날짜 전용: 품절 + 복원 날짜 모두
    # =========================
    if intent == "OUT_AND_RESTORE_DATES":
        res_out = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "OUT_OF_STOCK")
        )
        if date_from:
            res_out = res_out.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_out = res_out.lte("date", date_to.strftime("%Y-%m-%d"))
        res_out = res_out.execute()

        res_restore = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "RESTOCK")
        )
        if date_from:
            res_restore = res_restore.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_restore = res_restore.lte("date", date_to.strftime("%Y-%m-%d"))
        res_restore = res_restore.execute()

        if not res_out.data and not res_restore.data:
            return "해당 기간 내 품절 또는 복원 이력이 없습니다."

        out_map = {}
        for r in (res_out.data or []):
            key = str(r["product_url"]).strip().lower()
            out_map.setdefault(key, []).append(r["date"])
        restore_map = {}
        for r in (res_restore.data or []):
            key = str(r["product_url"]).strip().lower()
            restore_map.setdefault(key, []).append(r["date"])

        all_urls = set(list(out_map.keys()) + list(restore_map.keys()))
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(all_urls)]
        if df.empty:
            return "해당 제품의 품절 또는 복원 이력이 없습니다."

        results = []
        product_details = {}
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            out_dates = sorted(out_map.get(url, []))
            restore_dates = sorted(restore_map.get(url, []))
            all_events = (
                [(d, "❌ 품절") for d in out_dates] +
                [(d, "🔄 복원") for d in restore_dates]
            )
            all_events.sort(key=lambda x: x[0])
            timeline_str = " → ".join([f"{label} {d}" for d, label in all_events])
            results.append({
                "text": f"• {row['brand']} - {row['product_name']}\n  {timeline_str}",
                "product_url": url
            })
            product_details[url] = timeline_str

        return {
            "type": "product_list",
            "text": f"품절 및 복원 날짜 정보 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }

    if intent == "NEW":
        res = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "NEW_PRODUCT")
        )
        if date_from:
            res = res.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res = res.lte("date", date_to.strftime("%Y-%m-%d"))
        res = res.execute()
        if not res.data:
            return None
        new_product_data = {str(r["product_url"]).strip().lower(): r["date"] for r in res.data}

        
        df = df_work[
            df_work["product_url"].str.strip().str.lower().isin(new_product_data.keys())
        ].copy()
        
        if df.empty:
            return None
        
        df["product_url_key"] = df["product_url"].astype(str).str.strip().str.lower()
        df["launch_date"] = pd.to_datetime(df["product_url_key"].map(new_product_data))
        
        if any(k in question for k in ["순서","최신","최근"]):
            df = df.sort_values("launch_date", ascending=False)
        else:
            df = df.sort_values(["brand","category1","category2","product_name"])
        
        results = []
        product_details = {}
        
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            launch_date = row["launch_date"]
            categories = []
            if pd.notna(row.get("category1")) and row["category1"]:
                categories.append(row["category1"])
            if pd.notna(row.get("category2")) and row["category2"]:
                categories.append(row["category2"])
            category_str = f" [{' > '.join(categories)}]" if categories else ""
            product_name = row['product_name']
            url = str(row["product_url"]).strip().lower()
            launch_date = new_product_data.get(url)
            results.append({
                "text": f"• {row['brand']} - {product_name}{category_str}\n  🆕 출시일: {launch_date}",
                "product_url": url
            })
            product_details[url] = f"🆕 출시일: {launch_date}"
        if not results:
            return None
        return {
            "type": "product_list",
            "text": f"{period_label} 신제품 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
            "launch_dates": new_product_data
        }

    if intent == "OUT":
        date_from_str = date_from.strftime("%Y-%m-%d") if date_from else None
        date_to_str = date_to.strftime("%Y-%m-%d") if date_to else None

        # 1) product_lifecycle_events 에서 OUT_OF_STOCK / RESTOCK 수집
        res_lc = supabase.table("product_lifecycle_events").select("product_url, date, lifecycle_event")
        if date_from_str:
            res_lc = res_lc.gte("date", date_from_str)
        if date_to_str:
            res_lc = res_lc.lte("date", date_to_str)
        res_lc = res_lc.execute()

        out_map = {}
        restore_map = {}
        for r in (res_lc.data or []):
            key = str(r["product_url"]).strip().lower()
            if r["lifecycle_event"] == "OUT_OF_STOCK":
                out_map.setdefault(key, set()).add(r["date"])
            elif r["lifecycle_event"] == "RESTOCK":
                restore_map.setdefault(key, set()).add(r["date"])

        # 2) product_price_change_events 에서 추가 품절(unit_price=0) / 복원(prev_price=0) 수집
        res_pc = supabase.table("product_price_change_events").select("product_url, date, unit_price, prev_price")
        if date_from_str:
            res_pc = res_pc.gte("date", date_from_str)
        if date_to_str:
            res_pc = res_pc.lte("date", date_to_str)
        res_pc = res_pc.execute()

        for r in (res_pc.data or []):
            key = str(r["product_url"]).strip().lower()
            try:
                cur = float(r["unit_price"]) if r["unit_price"] is not None else None
                prv = float(r["prev_price"]) if r["prev_price"] is not None else None
            except (TypeError, ValueError):
                continue
            if cur == 0 and prv and prv > 0:
                out_map.setdefault(key, set()).add(r["date"])
            elif prv == 0 and cur and cur > 0:
                restore_map.setdefault(key, set()).add(r["date"])

        if not out_map:
            return None

        # set → sorted list
        out_map = {k: sorted(v) for k, v in out_map.items()}
        restore_map = {k: sorted(v) for k, v in restore_map.items()}

        urls = list(out_map.keys())
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(urls)].drop_duplicates(subset=["product_url"])
        if df.empty:
            return None

        results = []
        product_details = {}
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            out_dates = out_map.get(url, [])
            restore_dates = restore_map.get(url, [])

            all_events = (
                [(d, "❌ 품절") for d in out_dates] +
                [(d, "🔄 복원") for d in restore_dates]
            )
            all_events.sort(key=lambda x: x[0])
            timeline_str = " → ".join([f"{label} {d}" for d, label in all_events])

            categories = []
            if pd.notna(row.get("category1")) and row["category1"]:
                categories.append(row["category1"])
            if pd.notna(row.get("category2")) and row["category2"]:
                categories.append(row["category2"])
            category_str = f" [{' > '.join(categories)}]" if categories else ""

            results.append({
                "text": f"• {row['brand']} - {row['product_name']}{category_str}\n  {timeline_str}",
                "product_url": url
            })
            product_details[url] = timeline_str

        if not results:
            return None
        # 제품 수 기준 (품절 횟수 아님)
        unique_product_count = len(set(r["product_url"] for r in results))
        return {
            "type": "product_list",
            "text": f"{period_label} 품절 제품 ({unique_product_count}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }

    if intent == "RESTORE":
        res_restore = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "RESTOCK")
        )
        if date_from:
            res_restore = res_restore.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_restore = res_restore.lte("date", date_to.strftime("%Y-%m-%d"))
        res_restore = res_restore.execute()
        if not res_restore.data:
            return None

        res_out = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "OUT_OF_STOCK")
        )
        if date_from:
            res_out = res_out.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_out = res_out.lte("date", date_to.strftime("%Y-%m-%d"))
        res_out = res_out.execute()

        restore_map = {}
        for r in res_restore.data:
            key = str(r["product_url"]).strip().lower()
            restore_map.setdefault(key, []).append(r["date"])
        out_map = {}
        for r in (res_out.data or []):
            key = str(r["product_url"]).strip().lower()
            out_map.setdefault(key, []).append(r["date"])

        urls = list(restore_map.keys())
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(urls)]
        if df.empty:
            return None

        results = []
        product_details = {}
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            out_dates = sorted(out_map.get(url, []))
            restore_dates = sorted(restore_map.get(url, []))
            all_events = (
                [(d, "❌ 품절") for d in out_dates] +
                [(d, "🔄 복원") for d in restore_dates]
            )
            all_events.sort(key=lambda x: x[0])
            timeline_str = " → ".join([f"{label} {d}" for d, label in all_events])
            categories = []
            if pd.notna(row.get("category1")) and row["category1"]:
                categories.append(row["category1"])
            if pd.notna(row.get("category2")) and row["category2"]:
                categories.append(row["category2"])
            category_str = f" [{' > '.join(categories)}]" if categories else ""
            product_name = row['product_name']
            results.append({
                "text": f"• {row['brand']} - {product_name}{category_str}\n  {timeline_str}",
                "product_url": url
            })
            product_details[url] = timeline_str

        if not results:
            return None
        return {
            "type": "product_list",
            "text": f"{period_label} 복원 제품 ({len(results)}개)",
            "products": [
                str(r["product_url"]).strip().lower()
                for r in results
                if r.get("product_url")
            ],
            "product_details": product_details,
        }

    if intent == "OUT_AND_RESTORE":
        res_out = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "OUT_OF_STOCK")
        )
        if date_from:
            res_out = res_out.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_out = res_out.lte("date", date_to.strftime("%Y-%m-%d"))
        res_out = res_out.execute()

        res_restore = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "RESTOCK")
        )
        if date_from:
            res_restore = res_restore.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_restore = res_restore.lte("date", date_to.strftime("%Y-%m-%d"))
        res_restore = res_restore.execute()

        # 리스트 맵 (복수 날짜 지원) - URL 정규화
        out_map = {}
        for r in (res_out.data or []):
            key = str(r["product_url"]).strip().lower()
            out_map.setdefault(key, []).append(r["date"])
        restore_map = {}
        for r in (res_restore.data or []):
            key = str(r["product_url"]).strip().lower()
            restore_map.setdefault(key, []).append(r["date"])

        all_urls = set(list(out_map.keys()) + list(restore_map.keys()))
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(all_urls)]
        if df.empty:
            return "해당 기간 품절 또는 복원 제품이 없습니다."

        results = []
        product_details = {}
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            out_dates = sorted(out_map.get(url, []))
            restore_dates = sorted(restore_map.get(url, []))
            # 시간순 인터리브
            all_events = (
                [(d, "❌ 품절") for d in out_dates] +
                [(d, "🔄 복원") for d in restore_dates]
            )
            all_events.sort(key=lambda x: x[0])
            timeline_str = " → ".join([f"{label} {d}" for d, label in all_events])
            categories = []
            if pd.notna(row.get("category1")) and row["category1"]:
                categories.append(row["category1"])
            if pd.notna(row.get("category2")) and row["category2"]:
                categories.append(row["category2"])
            category_str = f" [{' > '.join(categories)}]" if categories else ""
            results.append({
                "text": f"• {row['brand']} - {row['product_name']}{category_str}\n  {timeline_str}",
                "product_url": url
            })
            product_details[url] = timeline_str

        return {
            "type": "product_list",
            "text": f"{period_label} 품절/복원 제품 ({len(results)}개)",
            "products": [
                str(r["product_url"]).strip().lower()
                for r in results
                if r.get("product_url")
            ],
            "product_details": product_details,
        }

    if intent == "NEW_AND_OUT":
        res_new = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "NEW_PRODUCT")
        )
        if date_from:
            res_new = res_new.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_new = res_new.lte("date", date_to.strftime("%Y-%m-%d"))
        res_new = res_new.execute()
        new_data = {str(r["product_url"]).strip().lower(): r["date"] for r in (res_new.data or [])}

        res_out = (
            supabase.table("product_lifecycle_events")
            .select("product_url, date")
            .eq("lifecycle_event", "OUT_OF_STOCK")
        )
        if date_from:
            res_out = res_out.gte("date", date_from.strftime("%Y-%m-%d"))
        if date_to:
            res_out = res_out.lte("date", date_to.strftime("%Y-%m-%d"))
        res_out = res_out.execute()
        out_data = {str(r["product_url"]).strip().lower(): r["date"] for r in (res_out.data or [])}

        all_urls = set(list(new_data.keys()) + list(out_data.keys()))
        df = df_work[df_work["product_url"].str.strip().str.lower().isin(all_urls)]
        if df.empty:
            return "해당 기간 신제품 또는 품절 제품이 없습니다."

        new_results = []
        product_details = {}  # 🔥 추가
        for _, row in df[df["product_url"].str.strip().str.lower().isin(new_data)].iterrows():
            categories = []
            if pd.notna(row.get("category1")) and row["category1"]:
                categories.append(row["category1"])
            if pd.notna(row.get("category2")) and row["category2"]:
                categories.append(row["category2"])
            category_str = f" [{' > '.join(categories)}]" if categories else ""
            url = str(row["product_url"]).strip().lower()
            new_results.append({
                "text": f"• {row['brand']} - {row['product_name']}{category_str}\n  🆕 출시일: {new_data.get(url)}",
                "product_url": url
            })
            product_details[url] = f"🆕 출시일: {new_data.get(url)}"

        out_results = []
        for _, row in df[df["product_url"].str.strip().str.lower().isin(out_data)].iterrows():
            categories = []
            if pd.notna(row.get("category1")) and row["category1"]:
                categories.append(row["category1"])
            if pd.notna(row.get("category2")) and row["category2"]:
                categories.append(row["category2"])
            category_str = f" [{' > '.join(categories)}]" if categories else ""
            url = str(row["product_url"]).strip().lower()
            out_results.append({
                "text": f"• {row['brand']} - {row['product_name']}{category_str}\n  📅 품절일: {out_data.get(url)}",
                "product_url": url
            })
            product_details[url] = f"📅 품절일: {out_data.get(url)}"

        all_results = new_results + out_results

        if new_results and out_results:
            text = f"🆕 신제품 ({len(new_results)}개) + ❌ 품절 ({len(out_results)}개)"
        elif new_results:
            text = f"🆕 신제품 ({len(new_results)}개)"
        elif out_results:
            text = f"❌ 품절 제품 ({len(out_results)}개)"
        else:
            text = ""

        return {
            "type": "product_list",
            "text": text,
            "products": [
                str(r["product_url"]).strip().lower()
                for r in all_results
                if r.get("product_url")
            ],
            "new_products": [str(r["product_url"]).strip().lower() for r in new_results],
            "out_products": [str(r["product_url"]).strip().lower() for r in out_results],
            "product_details": product_details,  # 🔥 추가
        }

    if intent == "NORMAL_CHANGE":

        query = (
            supabase.table("product_normal_price_events")
            .select("product_url,date,prev_price,normal_price,price_diff")
            .gte("date", date_from.strftime("%Y-%m-%d"))
            .lte("date", date_to.strftime("%Y-%m-%d"))
        )
    
        res = query.order("date", desc=True).execute()
    
        if not res.data:
            return "해당 기간 내 정상가 변동이 없습니다."
    
        df = pd.DataFrame(res.data)
    
        product_details = {}
        results = []
        for _, row in df.iterrows():
            url = str(row["product_url"]).strip().lower()
            # ✅ df_summary 대신 df_work 사용 (브랜드 필터 반영)
            product_row = df_work[df_work["product_url"].str.strip().str.lower() == url]
            if product_row.empty:
                continue
            diff = float(row["price_diff"])
            arrow = "📈" if diff > 0 else "📉"
            detail = f"{arrow} 정상가 {float(row['prev_price']):,.0f}원 → {float(row['normal_price']):,.0f}원 ({diff:+,.0f}원) | {row['date']}"
            if url in product_details:
                product_details[url] += f"  /  {detail}"
            else:
                product_details[url] = detail
                results.append({"product_url": url})
        if not results:
            return "해당 기간 내 정상가 변동이 없습니다."
        return {
            "type": "product_list",
            "text": f"{period_label} 정상가 변동 제품 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }
    
    if intent == "VOLATILITY":

        res = (
            supabase.table("product_all_events")
            .select("product_url, unit_price, date")
            .in_("product_url", df_work["product_url"].tolist())   # 🔥 브랜드 필터 유지
            .gte("date", date_from.strftime("%Y-%m-%d"))
            .lte("date", date_to.strftime("%Y-%m-%d"))
            .execute()
        )
        if not res.data:
            return None
        df = pd.DataFrame(res.data)
        df["unit_price"] = df["unit_price"].astype(float)
        df["date"] = pd.to_datetime(df["date"])
        
        # 🔥 조회기간 필터
        df = df[
            (df["date"] >= pd.Timestamp(date_from)) &
            (df["date"] <= pd.Timestamp(date_to))
        ]
        volatility = (
            df.groupby("product_url")["unit_price"]
            .agg(lambda x: x.max() - x.min())
            .sort_values(ascending=False)
        )
        if volatility.empty:
            return None
        product_details = {}
        results = []
        for url, val in volatility.items():
            if val == 0:
                continue
            norm_url = str(url).strip().lower()
            row = df_work[df_work["product_url"].str.strip().str.lower() == norm_url]
            if row.empty:
                continue
            lo = df[df["product_url"] == url]["unit_price"].min()
            hi = df[df["product_url"] == url]["unit_price"].max()
            product_details[norm_url] = f"💰 최저 {lo:,.1f}원 ~ 최고 {hi:,.1f}원 (변동폭 {val:,.1f}원)"
            results.append({"product_url": norm_url})
        if not results:
            return None
        return {
            "type": "product_list",
            "text": f"{period_label} 가격 변동 제품 ({len(results)}개)",
            "products": [r["product_url"] for r in results],
            "product_details": product_details,
        }


    # =========================
    # 🔍 UNKNOWN: 키워드 제품 검색
    # =========================
    # intent가 분류되지 않았거나, 키워드로 제품을 찾으려는 경우
    keywords = [w for w in question.split() if len(w) >= 2]
    if keywords:
        df_search = df_work.copy()
        for keyword in keywords:
            nkw = _norm_kw(keyword)
            mask = (
                _norm_series(df_search["product_name"]).str.contains(nkw, case=False) |
                _norm_series(df_search["brand"]).str.contains(nkw, case=False) |
                _norm_series(df_search["category1"]).str.contains(nkw, case=False) |
                _norm_series(df_search["category2"]).str.contains(nkw, case=False) |
                _norm_series(df_search["brew_type_kr"]).str.contains(nkw, case=False)
            )
            df_search = df_search[mask]
            if df_search.empty:
                break

        df_search = df_search.drop_duplicates(subset=["product_url"])
        if not df_search.empty:
            products = [str(r["product_url"]).strip().lower() for _, r in df_search.iterrows()]
            return {
                "type": "product_list",
                "text": f"'{question}' 검색 결과 ({len(products)}개)",
                "products": products,
            }

    return None

def llm_fallback(question: str, df_summary: pd.DataFrame):
    context = df_summary.head(50)[
        ["product_name", "current_unit_price", "is_discount", "is_new_product", "brew_type_kr"]
    ].to_dict(orient="records")

    prompt = f"""
당신은 커피 캡슐 가격 분석 전문가입니다.
아래 데이터 기반으로 질문에 답하세요.

규칙:
- 답변은 간결하게 핵심만 말하세요.
- 데이터 컬럼명(is_new_product, is_discount 등 영문 필드명)은 절대 언급하지 마세요.
- 기술적인 내부 정보(컬럼명, 값, 코드)는 노출하지 마세요.
- 결과가 없으면 "해당 조건에 맞는 제품이 없습니다." 로만 답하세요.

데이터:
{context}

질문:
{question}
"""

    from openai import OpenAI
    client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    return response.choices[0].message.content

# =========================
# 3️⃣ 유틸 (제품명 보정 포함)
# =========================

import re


def clean_product_name(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = re.sub(r"[\x00-\x1f\x7f-\x9f]", "", s).strip()
    fix_map = {
        "본   직영": "본사직영",
        "본  직영": "본사직영",
        "본 직영": "본사직영",
        "바닐   향": "바닐라향",
        "바닐  향": "바닐라향",
        "네스프   ": "네스프레소",
        "스타   스": "스타벅스",
    }
    for bad, good in fix_map.items():
        if bad in s:
            s = s.replace(bad, good)
    s = re.sub(r"바닐.*?향", "바닐라향", s)
    s = re.sub(r"본.*?직영", "본사직영", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s

def detect_encoding_issues(df: pd.DataFrame):
    if "product_name_raw" not in df.columns:
        return
    mask = df["product_name_raw"].str.contains(" ", na=False)
    issues = df[mask][["product_url", "product_name_raw"]]
    if not issues.empty:
        import logging
        logging.warning(f"[ENCODING ISSUE] {len(issues)}건 감지됨")
        try:
            supabase.table("product_name_encoding_issues").insert(
                issues.to_dict(orient="records")
            ).execute()
        except Exception as e:
            logging.error(f"로그 저장 실패: {e}")


def _norm_series(s: pd.Series) -> pd.Series:
    return (
        s.fillna("")
        .astype(str)
        .str.lower()
        .str.replace(" ", "", regex=False)
    )

# ✅ 이 줄 추가
def _norm_kw(kw: str) -> str:
    return kw.lower().replace(" ", "")

def options_from(df: pd.DataFrame, col: str):
    if col not in df.columns:
        return []

    vals = (
        df[col]
        .dropna()
        .astype(str)
        .str.strip()
    )

    vals = vals[~vals.isin(["None", "nan", ""])]

    return sorted(vals.unique().tolist())

def format_product_label(row):
    brand = row.get("brand")
    product_name = row.get("product_name")
    category1 = row.get("category1")
    category2 = row.get("category2")

    brand = str(brand).strip() if pd.notna(brand) else ""
    product_name = str(product_name).strip() if pd.notna(product_name) else ""
    category1 = str(category1).strip() if pd.notna(category1) else ""
    category2 = str(category2).strip() if pd.notna(category2) else ""

    if brand in {"카누 바리스타", "네슬레", "일리카페"}:
        return f"{brand} - {product_name}"

    if brand == "네스프레소":
        parts = [brand]
        if category2:
            parts.append(category2)
        parts.append(product_name)
        return " - ".join(parts)

    parts = [brand]
    if category1:
        parts.append(category1)
    if category2:
        parts.append(category2)
    parts.append(product_name)
    return " - ".join(parts)


import re

def render_card(bg, border, title, content):
    return f"""
    <div style="
        background:{bg};
        padding:18px;
        border-radius:12px;
        border-left:6px solid {border};
        min-height:120px;
        box-shadow:0 1px 3px rgba(0,0,0,0.06);
    ">
        <div style="font-weight:600;font-size:15px;margin-bottom:8px;">
            {title}
        </div>
        {content}
    </div>
    """

# =========================
# 🔧 제품 선택 토글 함수 (안정화)
# =========================
def toggle_product(product_url):
    if product_url in st.session_state.selected_products:
        st.session_state.selected_products.remove(product_url)
    else:
        st.session_state.selected_products.add(product_url)

# =========================
# 정상가 변동
# =========================
def get_normal_price_change_dates(product_url, date_from, date_to):
    res = (
        supabase.table("product_normal_price_events")
        .select("date")
        .eq("product_url", product_url)
        .gte("date", date_from.strftime("%Y-%m-%d"))
        .lte("date", date_to.strftime("%Y-%m-%d"))
        .execute()
    )
    if not res.data:
        return []
    return [r["date"] for r in res.data]

# =========================
# 4️⃣ 체크박스 key 등록
# =========================
def register_product_checkbox_key(product_url: str, widget_key: str):
    if "product_checkbox_keys" not in st.session_state:
        st.session_state["product_checkbox_keys"] = {}
    st.session_state["product_checkbox_keys"].setdefault(product_url, set()).add(widget_key)

def remove_product_everywhere(product_url: str):
    clean_url = str(product_url).strip("_").strip()
    st.session_state.selected_products.discard(clean_url)

# =========================
# 4️⃣ 세션 상태 초기화
# =========================
if "selected_products" not in st.session_state:
    st.session_state.selected_products = set()

if "keyword_results" not in st.session_state:
    st.session_state.keyword_results = {}
if "active_mode" not in st.session_state:
    st.session_state.active_mode = "키워드 검색"
if "show_results" not in st.session_state:
    st.session_state.show_results = False
if "search_keyword" not in st.session_state:
    st.session_state.search_keyword = ""
if "search_history" not in st.session_state:
    st.session_state.search_history = []

if "selected_products" not in st.session_state:
    st.session_state.selected_products = set()

if "_removed_products" not in st.session_state:
    st.session_state["_removed_products"] = set()

# =========================
# 5️⃣ 메인 UI
# =========================

# =====개발용 환경 여기만 다름====================
# -------------------------
# 🔐 비밀번호 인증
# -------------------------
def check_password():
    app_password = st.secrets.get("APP_PASSWORD", "")
    if not app_password:
        return True  # secrets에 설정 안 된 경우 통과
    if st.session_state.get("authenticated"):
        return True
    st.title("☕ Coffee Capsule Price Intelligence")
    st.markdown("---")
    col1, col2, col3 = st.columns([3, 2, 3])
    with col2:
        st.markdown("### 🔒 로그인")
        pw = st.text_input("비밀번호를 입력하세요", type="password", key="pw_input")
        if st.button("로그인", use_container_width=True):
            if pw == app_password:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("비밀번호가 틀렸습니다.")
    return False

if not check_password():
    st.stop()

# =====개발용 환경 여기만 다름====================

st.title("☕ Coffee Capsule Price Intelligence")


# -------------------------

# 데이터 로딩 (탭 이전에 로드)

# -------------------------

df_all = load_product_summary()


# -------------------------
# 브랜드 정규화
# -------------------------
df_all["brand"] = (
    df_all["brand"]
    .astype(str)
    .str.strip()
    .str.replace("�", "", regex=False)
)

df_all["brand"] = df_all["brand"].replace({
    "네레": "네슬레",
    "네스프레": "네스프레소",
    "일리": "일리카페",
    "카누": "카누 바리스타",
    "카누 바스타": "카누 바리스타"
})

# -------------------------
# category1 정규화
# -------------------------
df_all["category1"] = (
    df_all["category1"]
    .astype(str)
    .str.strip()
    .str.replace("�", "", regex=False)
)

# 카누 전용
df_all["category1"] = df_all["category1"].str.replace(
    r"카.*바리스타.*캡슐",
    "카누 바리스타 전용캡슐",
    regex=True
)

# 카누 네스프레소 호환
df_all["category1"] = df_all["category1"].str.replace(
    r"카.*네스프레소.*캡슐",
    "카누 네스프레소 호환캡슐",
    regex=True
)

# 카누 돌체구스토 호환
df_all["category1"] = df_all["category1"].str.replace(
    r"카.*돌체.*캡슐",
    "카누 돌체구스토 호환캡슐",
    regex=True
)

# 돌체구스토 통합
df_all["category1"] = df_all["category1"].str.replace(
    r"돌체.*캡.*",
    "돌체구스토 캡슐",
    regex=True
)

# 스타벅스 오타 수정
df_all["category1"] = df_all["category1"].replace({
    "스타벅스by네스프소": "스타벅스by네스프레소"
})

df_all["category2"] = (
    df_all["category2"]
    .astype(str)
    .str.strip()
    .str.replace("�", "", regex=False)
)

df_all["category2"] = df_all["category2"].replace({
    "버츄": "버츄오",
    "버추오": "버츄오",
    "오리": "오리지널"
})

df_all["category2"] = df_all["category2"].replace({
    "None": None
})

# -------------------------
# 🔎 검색용 컬럼 생성 (공백 무시 검색)
# -------------------------

df_all["product_name_search"] = (
    df_all["product_name"]
    .astype(str)
    .str.lower()
    .str.replace(" ", "", regex=False)
)

df_all["brew_type_search"] = (
    df_all["brew_type_kr"]
    .astype(str)
    .str.lower()
    .str.replace(" ", "", regex=False)
)

df_all["category1_search"] = (
    df_all["category1"]
    .astype(str)
    .str.lower()
    .str.replace(" ", "", regex=False)
)

df_all["category2_search"] = (
    df_all["category2"]
    .astype(str)
    .str.lower()
    .str.replace(" ", "", regex=False)
)

# -------------------------
# URL 정리
# -------------------------
df_all["product_url"] = (
    df_all["product_url"]
    .astype(str)
    .str.strip("_")
    .str.strip()
)

if df_all is None or df_all.empty:
    st.warning("아직 집계된 제품 데이터가 없습니다.")
    st.stop()

# 제품명 정제
df_all["product_name_raw"] = df_all["product_name"]
df_all["product_name"] = df_all["product_name"].apply(clean_product_name)

# 🔎 검색용 컬럼 생성
df_all["product_name_search"] = (
    df_all["product_name"]
    .astype(str)
    .str.lower()
    .str.replace(" ", "", regex=False)
)

# -------------------------
# 깨진 문자열 감지 (운영 로그 전용)
# -------------------------
try:
    encoding_issues = detect_encoding_issues(df_all)
    if isinstance(encoding_issues, pd.DataFrame) and not encoding_issues.empty:
        print(f"[ENCODING] 깨진 제품명 {len(encoding_issues)}건 감지")
        log_records = encoding_issues[[
            "product_url",
            "product_name_raw"
        ]].to_dict(orient="records")
        supabase.table("product_name_encoding_issues") \
                .insert(log_records) \
                .execute()
except Exception as e:
    print(f"[ENCODING_LOG_ERROR] {e}")

# -------------------------
# 조회 기준 선택 및 조회 조건 통합
# -------------------------
col_main_left, col_main_right = st.columns([3, 1])

with col_main_left:
    st.subheader("🔎 조회 기준")

with col_main_right:
    st.subheader("📅 조회 기간")

col_tabs, col_controls = st.columns([3, 1])
with col_controls:
    _min_date = pd.to_datetime(df_all["first_seen_date"].dropna()).min().date()
    _max_date = pd.to_datetime(df_all["last_seen_date"].dropna()).max().date()
    col_from, col_to = st.columns(2)
    with col_from:
        st.write("시작일")
        date_from = st.date_input(
            "시작일",
            value=_min_date,
            min_value=_min_date,
            max_value=_max_date,
            key="date_from",
            label_visibility="collapsed"
        )
    with col_to:
        st.write("종료일")
        date_to = st.date_input(
            "종료일",
            value=_max_date,
            min_value=_min_date,
            max_value=_max_date,
            key="date_to",
            label_visibility="collapsed"
        )

    st.button("📊 조회하기", type="primary", use_container_width=True, key="btn_search_trigger", on_click=lambda: st.session_state.update({"show_results": True}))

    if st.button("🗑️ 전체 초기화", use_container_width=True, key="btn_reset_all"):
        st.session_state.selected_products = set()
        st.session_state.keyword_results = {}
        st.session_state.show_results = False
        st.session_state.search_keyword = ""
        st.session_state.search_history = []

        if "insight_question" in st.session_state:
            del st.session_state.insight_question
        if "insight_question_input" in st.session_state:
            del st.session_state.insight_question_input
        if "question_history" in st.session_state:
            st.session_state.question_history = []

        st.session_state["filter_brand"] = "(전체)"
        st.session_state["filter_cat1"] = "(전체)"
        st.session_state["filter_cat2"] = "(전체)"
        if "last_filter" in st.session_state:
            del st.session_state.last_filter

        keys_to_delete = [
            key for key in list(st.session_state.keys())
            if key.startswith(("tab", "chk_tab", "remove_product_", "delete_"))
        ]
        st.session_state.selected_products = set()
        st.session_state["_removed_products"] = set()
        for key in keys_to_delete:
            del st.session_state[key]
        st.rerun()


with col_tabs:
    tab1, tab2, tab3 = st.tabs(["🔍 키워드 검색", "🎛️ 필터 선택", "🤖 자연어 질문"])

    # =========================
    # TAB 1: 키워드 검색
    # =========================
    with tab1:
        with st.form("search_form", clear_on_submit=True):
            keyword_input = st.text_input(
                "제품명 검색",
                placeholder="예: 카누 디카페인 (공백=AND) / 쥬시, 멜로지오 (쉼표=OR)",
                key="keyword_input_field"
            )
            submitted = st.form_submit_button("검색")

        if submitted and keyword_input.strip():
            search_keyword = keyword_input.strip()
            st.session_state.search_keyword = search_keyword
            st.session_state.active_mode = "키워드 검색"
            st.session_state["_auto_expand"] = True

            if "," in search_keyword:
                keywords = [k.strip() for k in search_keyword.split(",") if k.strip()]
                mask = False
                for kw in keywords:
                    nkw = _norm_kw(kw)
                    mask |= _norm_series(df_all["product_name"]).str.contains(nkw, case=False)
                    mask |= _norm_series(df_all["brand"]).str.contains(nkw, case=False)
                    mask |= _norm_series(df_all["category1"]).str.contains(nkw, case=False)
                    mask |= _norm_series(df_all["category2"]).str.contains(nkw, case=False)
                    mask |= _norm_series(df_all["brew_type_kr"]).str.contains(nkw, case=False)
                candidates_df = df_all[mask].copy()
            else:
                keywords = search_keyword.split()
                candidates_df = df_all.copy()
                for kw in keywords:
                    if len(kw) >= 2:
                        nkw = _norm_kw(kw)
                        keyword_mask = (
                            _norm_series(candidates_df["product_name"]).str.contains(nkw, case=False) |
                            _norm_series(candidates_df["brand"]).str.contains(nkw, case=False) |
                            _norm_series(candidates_df["category1"]).str.contains(nkw, case=False) |
                            _norm_series(candidates_df["category2"]).str.contains(nkw, case=False) |
                            _norm_series(candidates_df["brew_type_kr"]).str.contains(nkw, case=False)
                        )
                        candidates_df = candidates_df[keyword_mask]
                        if candidates_df.empty:
                            break

            try:
                supabase.table("search_logs").insert({
                    "search_type": "KEYWORD",
                    "search_term": search_keyword,
                    "result_count": len(candidates_df),
                    "created_at": datetime.now().isoformat()
                }).execute()
            except Exception as e:
                print("검색 로그 저장 실패:", e)

            existing_idx = None
            for idx, history in enumerate(st.session_state.search_history):
                if history["keyword"] == search_keyword:
                    existing_idx = idx
                    break

            sorted_df = (
                candidates_df
                .fillna("")
                .drop_duplicates(subset=["product_url"])
                .sort_values(by=["brand", "category1", "category2", "product_name"])
            )

            search_result = {
                "keyword": search_keyword,
                "results": sorted_df["product_url"].tolist()
            }

            if existing_idx is not None:
                st.session_state.search_history[existing_idx] = search_result
            else:
                st.session_state.search_history.append(search_result)

        st.markdown("### 📦 비교할 제품 선택")

        if not st.session_state.search_history:
            st.info("검색 결과가 없습니다.")
        else:
            num_cols = 3
            total_searches = len(st.session_state.search_history)

            for row_idx in range(0, total_searches, num_cols):
                cols = st.columns(num_cols)

                for col_idx in range(num_cols):
                    history_idx = row_idx + col_idx
                    if history_idx >= total_searches:
                        break

                    history = st.session_state.search_history[history_idx]

                    with cols[col_idx]:
                        with st.container(border=True):
                            col_title, col_delete = st.columns([4, 1])

                            with col_title:
                                st.markdown(f"**🔍 {history['keyword']}**")

                            with col_delete:
                                if st.button("🗑️", key=f"delete_search_{history_idx}", help="검색 결과 삭제"):
                                    for product_url in history["results"]:
                                        st.session_state.selected_products.discard(product_url)
                                        if "product_checkbox_keys" in st.session_state:
                                            keys = st.session_state["product_checkbox_keys"].get(product_url, set())
                                            for k in list(keys):
                                                if k in st.session_state:
                                                    del st.session_state[k]
                                    st.session_state.search_history.pop(history_idx)
                                    st.rerun()

                            if not history['results']:
                                st.caption("📭 검색 결과 없음")
                            else:
                                sorted_df = (
                                    df_all[df_all["product_url"].isin(history["results"])]
                                    .fillna("")
                                    .drop_duplicates(subset=["product_url"])
                                    .sort_values(by=["brand", "category1", "category2", "product_name"])
                                )

                                with st.expander(
                                    f"목록 펼치기 / 접기 ({len(sorted_df)}개)",
                                    expanded=st.session_state.get("_auto_expand", False)
                                ):
                                    for _, row in sorted_df.iterrows():
                                        product_url = row["product_url"]
                                        label = format_product_label(row)
                                        scope = f"hist_{history_idx}"
                                        is_selected = product_url in st.session_state.selected_products
                                        k = mk_widget_key("chk_tab1", product_url, scope) + ("_1" if is_selected else "_0")
                                        register_product_checkbox_key(product_url, k)
                                        col_chk, col_lbl = st.columns([0.06, 0.94], vertical_alignment="center")
                                        with col_chk:
                                            checked = st.checkbox("", key=k, value=is_selected)
                                        with col_lbl:
                                            st.markdown(
                                                f"""
                                                <div style="
                                                    white-space:normal;
                                                    word-break:keep-all;
                                                    overflow-wrap:break-word;
                                                    line-height:1.35;
                                                    padding:5px 0 6px 0;
                                                ">
                                                    {label}
                                                </div>
                                                """,
                                                unsafe_allow_html=True
                                            )
                                        if checked:
                                            st.session_state.selected_products.add(product_url)
                                        else:
                                            st.session_state.selected_products.discard(product_url)

                                    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

            if "_auto_expand" in st.session_state:
                st.session_state["_auto_expand"] = False

    # =========================
    # TAB 2: 필터 선택
    # =========================
    with tab2:
        col1, col2, col3 = st.columns(3)

        with col1:
            VALID_BRANDS = [
                "네스프레소",
                "네슬레",
                "일리카페",
                "카누 바리스타"
            ]
            
            brands = options_from(df_all, "brand")
            sel_brand = st.selectbox(
                "브랜드",
                ["(전체)"] + brands,
                index=0,
                key="filter_brand"
            )

        df1 = df_all if sel_brand == "(전체)" else df_all[df_all["brand"] == sel_brand]

        with col2:
            VALID_CAT1 = [
                "돌체구스토 캡슐",
                "스타벅스by네스프레소",
                "카누 네스프레소 호환캡슐",
                "카누 돌체구스토 호환캡슐",
                "카누 바리스타 전용캡슐",
                "캡슐",
                "커피"
            ]

            cat1s = options_from(df1, "category1")
            sel_cat1 = st.selectbox(
                "카테고리1",
                ["(전체)"] + cat1s,
                index=0,
                key="filter_cat1"
            )

        df2 = df1 if sel_cat1 == "(전체)" else df1[df1["category1"] == sel_cat1]

        with col3:
            VALID_CAT2 = ["버츄오", "오리지널"]

            cat2s = options_from(df2, "category2")
            sel_cat2 = st.selectbox(
                "카테고리2",
                ["(전체)"] + cat2s,
                index=0,
                key="filter_cat2"
            )

        candidates_df = df2 if sel_cat2 == "(전체)" else df2[df2["category2"] == sel_cat2]

        if sel_brand != "(전체)" or sel_cat1 != "(전체)" or sel_cat2 != "(전체)":
            st.session_state.active_mode = "필터 선택"
            current_filter = f"{sel_brand}|{sel_cat1}|{sel_cat2}"
            if "last_filter" not in st.session_state or st.session_state.last_filter != current_filter:
                try:
                    supabase.table("search_logs").insert({
                        "search_type": "FILTER",
                        "search_term": current_filter,
                        "filter_data": {
                            "brand": sel_brand,
                            "category1": sel_cat1,
                            "category2": sel_cat2
                        },
                        "result_count": len(candidates_df),
                        "created_at": datetime.now().isoformat()
                    }).execute()
                    st.session_state.last_filter = current_filter
                except Exception as e:
                    print("필터 로그 저장 실패:", e)

        st.markdown("### 📦 비교할 제품 선택")

        unique_df = (
            candidates_df
            .fillna("")
            .drop_duplicates(subset=["product_url"])
            .sort_values(by=["brand", "category1", "category2", "product_name"])
        )

        with st.expander(f"목록 펼치기 / 접기 ({len(unique_df)}개)", expanded=False):
            for _, row in unique_df.iterrows():
                product_url = row["product_url"]
                label = format_product_label(row)
                scope = f"{sel_brand}|{sel_cat1}|{sel_cat2}"
                is_selected = product_url in st.session_state.selected_products
                k = mk_widget_key("chk_tab2", product_url, scope) + ("_1" if is_selected else "_0")
                register_product_checkbox_key(product_url, k)
                col_chk, col_lbl = st.columns([0.02, 0.98], vertical_alignment="center")
                with col_chk:
                    checked = st.checkbox("", key=k, value=is_selected)
                with col_lbl:
                    st.markdown(
                        f"""
                        <div style="
                            white-space:normal;
                            word-break:keep-all;
                            overflow-wrap:break-word;
                            line-height:1.35;
                            padding:5px 0 6px 0;
                        ">
                            {label}
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                if checked:
                    st.session_state.selected_products.add(product_url)
                else:
                    st.session_state.selected_products.discard(product_url)

            st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

    # =========================
    # TAB 3: 자연어 질문
    # =========================
    with tab3:
        with st.form("question_form", clear_on_submit=True):
            question = st.text_input(
                "자연어로 질문하세요",
                placeholder="예: 카누 바리스타 쥬시 할인 기간 / 네스프레소 최저가 제품",
                key="insight_question_input"
            )
            ask_question = st.form_submit_button("🔍 질문하기", type="primary", use_container_width=True)

        if ask_question and question:
            st.session_state.active_mode = "자연어 질문"
            st.session_state.question_history = []
            intent = classify_intent(question)

            date_from = st.session_state.get("date_from", datetime.now() - timedelta(days=90))
            date_to = st.session_state.get("date_to", datetime.now())

            if not isinstance(date_from, datetime):
                date_from = datetime.combine(date_from, datetime.min.time()) if hasattr(date_from, 'year') else datetime.now() - timedelta(days=90)
            if not isinstance(date_to, datetime):
                date_to = datetime.combine(date_to, datetime.min.time()) if hasattr(date_to, 'year') else datetime.now()

            filtered_df = df_all.copy()

            _top_n, _top_dir, _top_mode = extract_top_n(question)
            _top_n_arg = (_top_n, _top_dir, _top_mode) if _top_n else None
            
            answer = execute_rule(
                intent,
                question,
                filtered_df,
                date_from,
                date_to,
                top_n=_top_n_arg
            )

            filter_info = {
                "date_from": date_from.strftime("%Y-%m-%d") if hasattr(date_from, 'strftime') else str(date_from),
                "date_to": date_to.strftime("%Y-%m-%d") if hasattr(date_to, 'strftime') else str(date_to),
                "total_products": len(filtered_df),
                "filtered": len(filtered_df) < len(df_all)
            }

            if answer:
                save_question_log(question, intent, False, answer, filter_info)
                st.session_state.question_history.append({
                    "question": question,
                    "answer": answer,
                    "intent": intent
                })
            else:
                with st.spinner("분석 중..."):
                    answer = llm_fallback(question, filtered_df)
                    answer = {"type": "text", "text": answer}
                save_question_log(question, intent, True, answer, filter_info)
                st.session_state.question_history.append({
                    "question": question,
                    "answer": answer,
                    "intent": intent
                })

            st.rerun()

        # 🔥 질문 이력 표시
        if "question_history" in st.session_state and st.session_state.question_history:
            st.markdown("---")

            for idx, history in enumerate(reversed(st.session_state.question_history)):
                with st.container(border=True):
                    col_q, col_del = st.columns([10, 1])

                    with col_q:
                        st.markdown(f"**Q:** {history['question']}")

                    with col_del:
                        if st.button("🗑️", key=f"delete_q_{idx}", help="질문 삭제"):
                            st.session_state.question_history.pop(
                                len(st.session_state.question_history) - 1 - idx
                            )
                            st.rerun()

                    answer_data = history["answer"]

                    with st.expander("💬 답변 펼치기 / 접기", expanded=True):

                        if isinstance(answer_data, dict) and answer_data.get("type") == "product_list":

                            header_text = answer_data['text'].split('\n')[0]
                            st.markdown(f"**A:** {header_text}")

                            if answer_data.get("products"):

                                st.markdown(
                                    "<div style='font-size:13px; color:#6b7280; margin:6px 0 8px 0;'>"
                                    "* 비교할 제품을 선택해 주세요"
                                    "</div>",
                                    unsafe_allow_html=True
                                )

                                # 🔥 product_details 가져오기
                                product_details = answer_data.get("product_details", {})

                                # 🔥 체크박스 렌더링 헬퍼 - product_details 파라미터 추가
                                def render_product_checkboxes(product_urls, scope_prefix, product_details=None):

                                    product_urls_set = set(str(u).strip().lower() for u in product_urls)
                                
                                    sorted_df = (
                                        df_all[df_all["product_url"].astype(str).str.strip().str.lower().isin(product_urls_set)]
                                        .fillna("")
                                        .drop_duplicates(subset=["product_url"])
                                    )
                                
                                    # 🔥 질문에 '순서/최신/최근' 있으면 날짜순
                                    if any(k in history["question"] for k in ["순서","최신","최근"]):
                                        if "launch_dates" in answer_data:
                                            sorted_df["product_url_key"] = sorted_df["product_url"].astype(str).str.strip().str.lower()
                                            sorted_df["launch_date"] = sorted_df["product_url_key"].map(answer_data["launch_dates"])
                                            sorted_df = sorted_df.sort_values("launch_date", ascending=False)
                                    else:
                                        sorted_df = sorted_df.sort_values(
                                            by=["brand","category1","category2","product_name"]
                                        )                            
                                    if sorted_df.empty:
                                        st.caption("⚠️ 매칭되는 제품이 없습니다.")
                                        return
                                    for _, row in sorted_df.iterrows():
                                        product_url = row["product_url"]
                                        product_url_key = str(product_url).strip().lower()  # 🔥 product_details 키 조회용
                                        label = format_product_label(row)
                                        scope = f"tab3_{idx}_{scope_prefix}"
                                        is_selected = product_url in st.session_state.selected_products
                                        k = mk_widget_key("chk_tab3", product_url, scope) + ("_1" if is_selected else "_0")
                                        register_product_checkbox_key(product_url, k)
                                        col_chk, col_lbl = st.columns([0.02, 0.98], vertical_alignment="center")
                                        with col_chk:
                                            checked = st.checkbox("", key=k, value=is_selected)
                                        with col_lbl:
                                            # 🔥 날짜 정보 추가
                                            detail_html = ""
                                            if product_details and product_url_key in product_details:
                                                detail_text = product_details[product_url_key]
                                                detail_html = (
                                                    f"<div style='font-size:12px; color:#6b7280; margin-top:4px; line-height:1.8;'>"
                                                    f"{detail_text}"
                                                    f"</div>"
                                                )
                                            st.markdown(
                                                f"<div style='white-space:normal; word-break:keep-all; overflow-wrap:break-word; line-height:1.35; padding:5px 0 6px 0;'>"
                                                f"{label}"
                                                f"{detail_html}"
                                                f"</div>",
                                                unsafe_allow_html=True
                                            )
                                        if checked:
                                            st.session_state.selected_products.add(product_url)
                                        else:
                                            st.session_state.selected_products.discard(product_url)
                                    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

                                # 🔥 NEW_AND_OUT: 신제품/품절 구분 표시
                                new_products = answer_data.get("new_products", [])
                                out_products = answer_data.get("out_products", [])

                                if new_products or out_products:
                                    if new_products:
                                        st.markdown("**🆕 신제품**")
                                        render_product_checkboxes(new_products, "new", product_details)
                                    if out_products:
                                        st.markdown("**❌ 품절 제품**")
                                        render_product_checkboxes(out_products, "out", product_details)
                                else:
                                    render_product_checkboxes(answer_data["products"], "all", product_details)

                            else:
                                st.caption("표시할 제품이 없습니다.")

                        elif isinstance(answer_data, dict):
                            st.markdown(f"**A:** {answer_data.get('text', str(answer_data))}")

                        else:
                            st.markdown(f"**A:** {answer_data}")

# =========================
# 8️⃣ 결과 표시
# =========================
selected_products = list(st.session_state.selected_products)

if selected_products:

    st.divider()

    col_title, col_download = st.columns([4, 1])
    with col_title:
        st.subheader(f"📊 조회 결과 ({len(selected_products)}개 제품)")
    with col_download:
        download_placeholder = st.empty()

    if date_from > date_to:
        st.error("❌ 시작일이 종료일보다 늦습니다. 기간을 다시 설정해주세요.")
        st.stop()

    st.info(f"📅 조회 기간: {date_from.strftime('%Y-%m-%d')} ~ {date_to.strftime('%Y-%m-%d')}")

    timeline_rows = []
    lifecycle_rows = []

    filter_date_from = pd.to_datetime(date_from)
    filter_date_to = pd.to_datetime(date_to)

    df_all_events = load_events_bulk(
        selected_products,
        filter_date_from,
        filter_date_to
    )

    df_lifecycle_all = load_lifecycle_bulk(
        selected_products,
        filter_date_from,
        filter_date_to
    )

    df_raw_unit_all = load_raw_unit_bulk(
        selected_products,
        filter_date_from,
        filter_date_to
    )

    for product_url in selected_products:
        product_row = df_all[df_all["product_url"] == product_url]

        if product_row.empty:
            st.session_state.selected_products.discard(product_url)
            continue

        row = product_row.iloc[0]
        pname = row["product_name"]

        df_price = df_all_events[
            df_all_events["product_url"] == row["product_url"]
        ].copy()
        if not df_price.empty:
            tmp = df_price.copy()
            if row['brand'] == '네스프레소':
                cat2 = str(row.get('category2') or '').strip()
                display_name = f"{row['brand']} - {cat2} - {pname}" if cat2 else f"{row['brand']} - {pname}"
            else:
                display_name = f"{row['brand']} - {pname}"
            tmp["product_name"] = display_name
            tmp["event_date"] = pd.to_datetime(tmp["date"])

            tmp = tmp.sort_values("event_date")

            all_dates = pd.date_range(
                start=tmp["event_date"].min(),
                end=tmp["event_date"].max(),
                freq="D"
            )

            tmp = tmp.set_index("event_date").reindex(all_dates).reset_index()
            tmp.rename(columns={"index": "event_date"}, inplace=True)

            tmp["unit_price"] = tmp["unit_price"].ffill()
            tmp["product_name"] = tmp["product_name"].ffill()

            tmp.loc[tmp["unit_price"] == 0, "unit_price"] = None

            tmp["normal_price"] = None
            tmp["discount_rate"] = None
            tmp["price_detail"] = ""

            df_raw_tmp = df_raw_unit_all[
                df_raw_unit_all["product_url"] == row["product_url"]
            ].copy()

            if not df_raw_tmp.empty:
                df_raw_tmp["date"] = pd.to_datetime(df_raw_tmp["date"])

                if "normal_price" in tmp.columns:
                    tmp.drop(columns=["normal_price"], inplace=True)

                tmp = tmp.merge(
                    df_raw_tmp[["date", "unit_normal_price"]],
                    left_on="event_date",
                    right_on="date",
                    how="left"
                )

                tmp.rename(columns={"unit_normal_price": "normal_price"}, inplace=True)
                tmp.drop(columns=["date"], inplace=True, errors="ignore")
                tmp = tmp.reset_index(drop=True)

                tmp["normal_price"] = pd.to_numeric(tmp["normal_price"], errors="coerce")
                tmp["unit_price"] = pd.to_numeric(tmp["unit_price"], errors="coerce")

            tmp["discount_rate"] = None
            tmp["is_discount"] = (tmp["event_type"] == "DISCOUNT").astype(bool)

            mask = (
                (tmp["is_discount"]) &
                (tmp["normal_price"].notna()) &
                (tmp["unit_price"].notna()) &
                (tmp["normal_price"] > 0)
            )

            tmp.loc[mask, "discount_rate"] = (
                (tmp.loc[mask, "normal_price"] - tmp.loc[mask, "unit_price"])
                / tmp.loc[mask, "normal_price"]
            ) * 100

            for idx2, price_row in tmp.iterrows():
                if pd.isna(price_row["unit_price"]):
                    tmp.at[idx2, "price_detail"] = "품절"
                elif price_row["is_discount"]:
                    if pd.notna(price_row["normal_price"]):
                        tmp.at[idx2, "price_detail"] = (
                            f"정상가: {price_row['normal_price']:,.1f}원 → "
                            f"할인가: {price_row['unit_price']:,.1f}원 "
                            f"({price_row['discount_rate']:.0f}% 할인)"
                        )
                    else:
                        tmp.at[idx2, "price_detail"] = f"할인가: {price_row['unit_price']:,.1f}원"
                else:
                    tmp.at[idx2, "price_detail"] = f"정상가: {price_row['unit_price']:,.1f}원"

            df_life = df_lifecycle_all[
                df_lifecycle_all["product_url"] == product_url
            ].copy()

            if not df_life.empty:
                df_life["date"] = pd.to_datetime(df_life["date"])
                out_dates = sorted(
                    df_life[df_life["lifecycle_event"]=="OUT_OF_STOCK"]["date"].dt.date
                )
                restore_dates = sorted(
                    df_life[df_life["lifecycle_event"]=="RESTOCK"]["date"].dt.date
                )

                from bisect import bisect_right

                out_dates_sorted = sorted(out_dates)
                restore_dates_sorted = sorted(restore_dates)

                for r2 in tmp.itertuples():
                    current_date = r2.event_date.date()
                    i_out = bisect_right(out_dates_sorted, current_date) - 1
                    last_out = out_dates_sorted[i_out] if i_out >= 0 else None
                    i_restore = bisect_right(restore_dates_sorted, current_date) - 1
                    last_restore = restore_dates_sorted[i_restore] if i_restore >= 0 else None
                    if last_out and (not last_restore or last_out > last_restore):
                        tmp.at[r2.Index, "unit_price"] = None

            tmp.loc[tmp["unit_price"].isna(), "price_detail"] = "품절"
            tmp["price_status"] = tmp["is_discount"].map({True: "💸 할인 중", False: "정상가"})
            tmp.loc[tmp["unit_price"].isna(), "price_status"] = "품절"

            tmp["product_url"] = row["product_url"]

            timeline_rows.append(
                tmp[["product_url", "product_name", "event_date", "unit_price", "price_status", "price_detail"]]
            )

            df_life = df_lifecycle_all[
                df_lifecycle_all["product_url"] == product_url
            ].copy()
            if not df_life.empty:
                df_life["date"] = pd.to_datetime(df_life["date"], errors="coerce")
                df_life = df_life.dropna(subset=["date"])

                lc_tmp = df_life.copy()
                if row['brand'] == '네스프레소':
                    cat2 = str(row.get('category2') or '').strip()
                    display_name = f"{row['brand']} - {cat2} - {pname}" if cat2 else f"{row['brand']} - {pname}"
                else:
                    display_name = f"{row['brand']} - {pname}"
                lc_tmp["product_name"] = display_name
                lc_tmp["event_date"] = pd.to_datetime(lc_tmp["date"])

                lc_final = lc_tmp.drop_duplicates(subset=["product_name", "event_date", "lifecycle_event"])

                lc_final = lc_final[
                    (lc_final["event_date"] >= filter_date_from) &
                    (lc_final["event_date"] <= filter_date_to)
                ]

                if not tmp.empty:
                    zero_dates = tmp[tmp["unit_price"].isna()]["event_date"].tolist()
                    existing_out = lc_final[lc_final["lifecycle_event"] == "OUT_OF_STOCK"]["event_date"].tolist()
                    existing_restock = lc_final[lc_final["lifecycle_event"] == "RESTOCK"]["event_date"].tolist()

                    new_rows = []
                    for zdate in sorted(zero_dates):
                        if zdate in existing_out:
                            continue
                        prior_restocks = [d for d in existing_restock if d <= zdate]
                        prior_outs = [d for d in existing_out + [r["event_date"] for r in new_rows] if d <= zdate]
                        if not prior_outs or (prior_restocks and max(prior_restocks) > max(prior_outs)):
                            new_rows.append({
                                "product_name": display_name,
                                "event_date": zdate,
                                "lifecycle_event": "OUT_OF_STOCK"
                            })
                            existing_out.append(zdate)

                    if new_rows:
                        lc_final = pd.concat([lc_final, pd.DataFrame(new_rows)], ignore_index=True)

                raw_lc_res = (
                    supabase.table("raw_daily_prices")
                    .select("date, normal_price")
                    .eq("product_url", row["product_url"])
                    .order("date", desc=False)
                    .execute()
                )
                if raw_lc_res.data:
                    raw_lc_df = pd.DataFrame(raw_lc_res.data)
                    raw_lc_df["normal_price"] = raw_lc_df["normal_price"].astype(float)
                    raw_lc_df["date"] = pd.to_datetime(raw_lc_df["date"])
                    raw_lc_df["prev_price"] = raw_lc_df["normal_price"].shift(1)

                    restock_from_raw = raw_lc_df[
                        (raw_lc_df["prev_price"] == 0) & (raw_lc_df["normal_price"] > 0)
                    ]["date"].tolist()

                    existing_restock_dates = lc_final[lc_final["lifecycle_event"] == "RESTOCK"]["event_date"].tolist()

                    for rdate in restock_from_raw:
                        if rdate >= filter_date_from and rdate <= filter_date_to:
                            if rdate not in existing_restock_dates:
                                lc_final = pd.concat([lc_final, pd.DataFrame([{
                                    "product_name": display_name,
                                    "event_date": rdate,
                                    "lifecycle_event": "RESTOCK"
                                }])], ignore_index=True)

                if not lc_final.empty:
                    lifecycle_rows.append(
                        lc_final[["product_name", "event_date", "lifecycle_event"]]
                    )

            zero_price_dates = tmp[tmp["unit_price"].isna() & (tmp["price_detail"] == "품절")]["event_date"].tolist()

            if zero_price_dates and not df_life.empty:
                existing_out_dates = df_life[df_life["lifecycle_event"] == "OUT_OF_STOCK"]["date"].tolist()
                for zdate in zero_price_dates:
                    if zdate not in existing_out_dates:
                        new_row = pd.DataFrame([{
                            "date": zdate,
                            "lifecycle_event": "OUT_OF_STOCK"
                        }])
                        df_life = pd.concat([df_life, new_row], ignore_index=True)

    # =========================
    # 8-1️⃣ 개당 가격 타임라인 비교 차트
    # =========================

    if timeline_rows:

        df_timeline = pd.concat(timeline_rows, ignore_index=True)
        df_timeline = df_timeline.sort_values(["product_name", "event_date"])
        df_timeline["unit_price"] = pd.to_numeric(df_timeline["unit_price"], errors="coerce")
        df_timeline["segment"] = (
            df_timeline["unit_price"].isna()
            .groupby(df_timeline["product_name"])
            .cumsum()
        )
        df_chart = df_timeline.dropna(subset=["unit_price"]).copy()

        df_chart["dup_rank"] = (
            df_chart.groupby(["event_date", "unit_price"])
            .cumcount()
        )
        df_chart["event_date_jitter"] = (
            df_chart["event_date"] +
            pd.to_timedelta(df_chart["dup_rank"] * 0.06, unit="D")
        )

        def get_or_create_color_map(keys: list) -> dict:
            palette = [
                "#4c78a8","#f58518","#e45756","#72b7b2","#54a24b",
                "#eeca3b","#b279a2","#ff9da6","#9d755d","#bab0ac",
                "#1f77b4","#ff7f0e","#2ca02c","#d62728","#9467bd",
                "#8c564b","#e377c2","#7f7f7f","#bcbd22","#17becf",
            ]
            if "color_map" not in st.session_state:
                st.session_state["color_map"] = {}
            cmap = st.session_state["color_map"]
            for k in sorted(set([str(x) for x in keys if str(x).strip()])):
                if k not in cmap:
                    cmap[k] = palette[len(cmap) % len(palette)]
            st.session_state["color_map"] = cmap
            return cmap

        def color_dot(hex_color: str) -> str:
            return (
                f"<span style='"
                f"color:{hex_color};"
                f"font-size:22px;"
                f"margin-right:12px;"
                f"display:inline-block;"
                f"vertical-align:middle;"
                f"'>●</span>"
            )

        col_chart, col_legend = st.columns([3, 1])

        color_map = get_or_create_color_map(df_chart["product_name"].unique().tolist())
        color_domain = list(color_map.keys())
        color_range = [color_map[k] for k in color_domain]

        with col_chart:
            show_overlap = st.toggle("겹친 제품 수 표시", value=False, key="toggle_overlap")

            base_line = (
                alt.Chart(df_chart)
                .mark_line()
                .encode(
                    x=alt.X("event_date:T", title="날짜", axis=alt.Axis(format="%m/%d")),
                    y=alt.Y("unit_price:Q", title="개당 가격 (원)"),
                    color=alt.Color(
                        "product_name:N",
                        scale=alt.Scale(domain=color_domain, range=color_range),
                        legend=None
                    ),
                    detail="segment:N",
                )
            )

            df_points = (
                df_chart.groupby(["event_date", "unit_price"])
                .agg(
                    product_names=("product_name", lambda x: "\n".join(sorted(set(x)))),
                    price_detail=("price_detail", lambda x: " / ".join(dict.fromkeys(x))),
                    price_status=("price_status", "first"),
                    count=("product_name", "count"),
                    product_name=("product_name", "first"),
                )
                .reset_index()
            )

            df_overlap = df_points[df_points["count"] > 1].copy()
            df_single = df_points[df_points["count"] == 1].copy()

            point_single = (
                alt.Chart(df_single)
                .mark_point(size=60, filled=True)
                .encode(
                    x=alt.X("event_date:T"),
                    y=alt.Y("unit_price:Q"),
                    color=alt.Color(
                        "product_name:N",
                        scale=alt.Scale(domain=color_domain, range=color_range),
                        legend=None
                    ),
                    tooltip=[
                        alt.Tooltip("product_names:N", title="제품"),
                        alt.Tooltip("event_date:T", title="날짜", format="%Y-%m-%d"),
                        alt.Tooltip("price_detail:N", title="가격 정보"),
                        alt.Tooltip("price_status:N", title="상태"),
                    ],
                )
            )

            layers = [base_line, point_single]

            if not df_overlap.empty:
                point_overlap = (
                    alt.Chart(df_overlap)
                    .mark_point(size=120, filled=True, color="red")
                    .encode(
                        x=alt.X("event_date:T"),
                        y=alt.Y("unit_price:Q"),
                        tooltip=[
                            alt.Tooltip("product_names:N", title="제품 (겹침)"),
                            alt.Tooltip("event_date:T", title="날짜", format="%Y-%m-%d"),
                            alt.Tooltip("price_detail:N", title="가격 정보"),
                            alt.Tooltip("count:Q", title="겹친 제품 수"),
                        ],
                    )
                )
                layers.append(point_overlap)

                if show_overlap:
                    text_overlap = (
                        alt.Chart(df_overlap)
                        .mark_text(dy=-10, fontSize=11, fontWeight="bold", color="red")
                        .encode(
                            x=alt.X("event_date:T"),
                            y=alt.Y("unit_price:Q"),
                            text=alt.Text("count:Q"),
                        )
                    )
                    layers.append(text_overlap)

            if lifecycle_rows:
                df_life_all = pd.concat(lifecycle_rows, ignore_index=True)

                icon_config = {
                    "NEW_PRODUCT": {"color": "green", "label": "NEW"},
                    "OUT_OF_STOCK": {"color": "red", "label": "품절"},
                    "RESTOCK": {"color": "orange", "label": "복원"},
                }

                for event_type, cfg in icon_config.items():
                    df_filtered = df_life_all[df_life_all["lifecycle_event"] == event_type].copy()
                    if df_filtered.empty:
                        continue

                    df_filtered = df_filtered.merge(
                        df_timeline[["product_name", "event_date", "unit_price", "price_detail"]],
                        on=["product_name", "event_date"],
                        how="left"
                    )

                    if event_type in ["OUT_OF_STOCK", "RESTOCK"]:
                        if event_type == "OUT_OF_STOCK":
                            for idx2, r2 in df_filtered.iterrows():
                                product_prices = df_timeline[
                                    (df_timeline["product_name"] == r2["product_name"]) &
                                    (df_timeline["event_date"] <= r2["event_date"]) &
                                    (df_timeline["unit_price"].notna())
                                ]
                                if not product_prices.empty:
                                    closest = product_prices.sort_values("event_date").iloc[-1]
                                    df_filtered.at[idx2, "unit_price"] = closest["unit_price"]
                            df_filtered["price_detail"] = "-"
                            df_filtered["price_status"] = "품절"

                        elif event_type == "RESTOCK":
                            for idx2, r2 in df_filtered.iterrows():
                                product_prices = df_timeline[
                                    (df_timeline["product_name"] == r2["product_name"]) &
                                    (df_timeline["unit_price"].notna())
                                ]
                                if not product_prices.empty:
                                    pp = product_prices.copy()
                                    pp["date_diff"] = (pp["event_date"] - r2["event_date"]).abs()
                                    closest = pp.sort_values("date_diff").iloc[0]
                                    df_filtered.at[idx2, "unit_price"] = closest["unit_price"]
                                    df_filtered.at[idx2, "price_detail"] = closest["price_detail"]
                    else:
                        for idx2, r2 in df_filtered[df_filtered["unit_price"].isna()].iterrows():
                            product_prices = df_timeline[
                                (df_timeline["product_name"] == r2["product_name"]) &
                                (df_timeline["unit_price"].notna())
                            ]
                            if not product_prices.empty:
                                pp = product_prices.copy()
                                pp["date_diff"] = (pp["event_date"] - r2["event_date"]).abs().dt.total_seconds()
                                closest = pp.nsmallest(1, "date_diff").iloc[0]
                                df_filtered.at[idx2, "unit_price"] = closest["unit_price"]
                                df_filtered.at[idx2, "price_detail"] = closest["price_detail"]

                    if df_filtered.empty:
                        continue

                    event_label_map = {
                        "NEW_PRODUCT": "신제품",
                        "OUT_OF_STOCK": "품절",
                        "RESTOCK": "복원",
                    }
                    df_filtered["event_label"] = df_filtered["lifecycle_event"].map(event_label_map).fillna(df_filtered["lifecycle_event"])

                    point_layer = (
                        alt.Chart(df_filtered)
                        .mark_point(size=150, shape="triangle-up", color=cfg["color"])
                        .encode(
                            x="event_date:T",
                            y="unit_price:Q",
                            tooltip=[
                                alt.Tooltip("product_name:N", title="제품"),
                                alt.Tooltip("event_date:T", title="날짜", format="%Y-%m-%d"),
                                alt.Tooltip("price_detail:N", title="가격 정보"),
                                alt.Tooltip("event_label:N", title="이벤트"),
                            ],
                        )
                    )

                    text_layer = (
                        alt.Chart(df_filtered)
                        .mark_text(dy=12, fontSize=11, fontWeight="bold", color=cfg["color"])
                        .encode(
                            x="event_date:T",
                            y="unit_price:Q",
                            text=alt.value(cfg["label"]),
                        )
                    )

                    layers.append(point_layer)
                    layers.append(text_layer)

            chart = alt.layer(*layers).properties(height=420).interactive()
            st.altair_chart(chart, use_container_width=True)

        with col_legend:
            st.markdown("#### 📋 제품 목록")

            unique_urls = sorted(df_chart["product_url"].unique())

            for product_url in unique_urls:
                product_row = df_all[df_all["product_url"] == product_url]
                if product_row.empty:
                    continue

                row = product_row.iloc[0]
                label = format_product_label(row)
                display_name = f"{row['brand']} - {row['product_name']}"
                hex_color = color_map.get(display_name, "#999999")

                col_btn, col_name = st.columns([1, 10])

                with col_btn:
                    if st.button("×", key=f"remove_product_{product_url}", help="차트에서 제거"):
                        clean_url = product_url.strip("_")
                        remove_product_everywhere(clean_url)
                        st.rerun()

                with col_name:
                    html = (
                        f"<div style='display:flex; align-items:center; gap:12px;'>"
                        f"<span style='color:{hex_color}; font-size:22px; line-height:1; flex:0 0 auto;'>●</span>"
                        f"<div style='white-space:normal; word-break:keep-all; overflow-wrap:break-word; line-height:1.4;'>"
                        f"<b>{label}</b>"
                        f"</div>"
                        f"</div>"
                    )
                    st.markdown(html, unsafe_allow_html=True)

        # =========================
        # 🔥 엑셀 다운로드
        # =========================
        with download_placeholder:

            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            excel_data = df_chart[[
                "product_url",
                "product_name",
                "event_date",
                "unit_price",
                "price_status"
            ]].copy()

            excel_data["brand"] = excel_data["product_name"].str.split(" - ").str[0]
            excel_data["product_name_only"] = excel_data["product_name"].str.split(" - ").str[1]

            excel_data["event_date_str"] = pd.to_datetime(
                excel_data["event_date"]
            ).dt.strftime("%Y-%m-%d")

            raw_res = (
                supabase.table("raw_daily_prices_unit")
                .select("product_url, date, unit_normal_price")
                .in_("product_url", selected_products)
                .gte("date", filter_date_from.strftime("%Y-%m-%d"))
                .lte("date", filter_date_to.strftime("%Y-%m-%d"))
                .execute()
            )

            raw_df = pd.DataFrame(raw_res.data) if raw_res.data else pd.DataFrame()

            if not raw_df.empty:
                raw_df["date"] = pd.to_datetime(raw_df["date"]).dt.strftime("%Y-%m-%d")
                excel_data = excel_data.merge(
                    raw_df,
                    left_on=["product_url", "event_date_str"],
                    right_on=["product_url", "date"],
                    how="left"
                )
                excel_data.rename(columns={"unit_normal_price": "normal_price"}, inplace=True)
            else:
                excel_data["normal_price"] = None

            excel_data["discount_price"] = None
            excel_data["discount_rate"] = None

            mask_discount = excel_data["price_status"] == "💸 할인 중"
            excel_data.loc[mask_discount, "discount_price"] = excel_data.loc[mask_discount, "unit_price"]

            excel_data["normal_price"] = pd.to_numeric(excel_data["normal_price"], errors="coerce")
            excel_data["discount_price"] = pd.to_numeric(excel_data["discount_price"], errors="coerce")

            mask_valid = (
                mask_discount &
                excel_data["normal_price"].notna() &
                (excel_data["normal_price"] > 0)
            )

            discount_rate_series = (
                (excel_data.loc[mask_valid, "normal_price"]
                 - excel_data.loc[mask_valid, "discount_price"])
                / excel_data.loc[mask_valid, "normal_price"]
            ) * 100

            excel_data.loc[mask_valid, "discount_rate"] = (
                discount_rate_series.round(1).map(lambda x: f"{x:.1f}%")
            )

            mask_normal = excel_data["price_status"] != "💸 할인 중"
            excel_data.loc[mask_normal, "normal_price"] = excel_data.loc[mask_normal, "unit_price"]

            excel_data = excel_data[[
                "brand",
                "product_name_only",
                "event_date_str",
                "price_status",
                "normal_price",
                "discount_price",
                "discount_rate"
            ]]

            excel_data.columns = [
                "브랜드", "제품명", "날짜", "이벤트", "정상가", "할인가", "할인율"
            ]

            excel_data["정상가"] = pd.to_numeric(excel_data["정상가"], errors="coerce").round(1)
            excel_data["할인가"] = pd.to_numeric(excel_data["할인가"], errors="coerce").round(1)

            output = BytesIO()

            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                excel_data.to_excel(writer, sheet_name="가격 데이터", index=False)

                ws = writer.sheets["가격 데이터"]

                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                col_normal = excel_data.columns.get_loc("정상가") + 1
                col_discount = excel_data.columns.get_loc("할인가") + 1
                for row in ws.iter_rows(min_row=2, min_col=col_normal, max_col=col_discount):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = '#,##0.0'

                for i, col_name in enumerate(excel_data.columns, start=1):
                    series_as_str = excel_data[col_name].astype(str).fillna("")
                    max_len = max([len(str(col_name))] + series_as_str.map(len).tolist())
                    width = min(max(max_len + 2, 10), 60)
                    ws.column_dimensions[get_column_letter(i)].width = width

            output.seek(0)

            st.download_button(
                label="📥 엑셀 다운로드",
                data=output.getvalue(),
                file_name=f"Coffee Capsule Price Intelligence_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    else:
        st.info("다운로드할 데이터가 없습니다.")

    # =========================
    # 8-2️⃣ 제품별 카드
    # =========================

    for product_url in selected_products:

        product_row = df_all[df_all["product_url"] == product_url]

        if product_row.empty:
            st.session_state.selected_products.discard(product_url)
            continue

        p = product_row.iloc[0]
        label = format_product_label(p)
        st.markdown(f"### {label}")

        df_life = load_lifecycle_events(p["product_url"])
        if not df_life.empty:
            df_life["date"] = pd.to_datetime(df_life["date"], errors="coerce")
            df_life = df_life.dropna(subset=["date"])
            df_life = df_life[
                (df_life["date"] >= pd.Timestamp(filter_date_from)) &
                (df_life["date"] <= pd.Timestamp(filter_date_to))
            ]

        normal_change_res = (
            supabase.table("product_normal_price_events")
            .select("date, prev_price, normal_price, price_diff")
            .eq("product_url", p["product_url"])
            .gte("date", filter_date_from.strftime("%Y-%m-%d"))
            .lte("date", filter_date_to.strftime("%Y-%m-%d"))
            .order("date", desc=True)
            .limit(1)
            .execute()
        )

        normal_change_rows = normal_change_res.data if normal_change_res.data else []

        c1, c2, c3, c4 = st.columns(4)

        with c1:
            normal_value = p.get("normal_unit_price")
            if normal_value is not None and pd.notna(normal_value):
                if float(normal_value) == 0:
                    st.metric("개당 정상가", "품절", delta="재고 없음")
                else:
                    st.metric("개당 정상가", f"{float(normal_value):,.1f}원")
            else:
                st.metric("개당 정상가", "-")

        cards = []
        discount_res = supabase.rpc(
            "get_discount_periods_in_range",
            {
                "p_product_url": p["product_url"],
                "p_date_from": filter_date_from.strftime("%Y-%m-%d"),
                "p_date_to": filter_date_to.strftime("%Y-%m-%d"),
            }
        ).execute()

        discount_rows = discount_res.data if discount_res.data else []

        if discount_rows:
            latest_discount = discount_rows[-1]
            cards.append(render_card(
                "#e9f3ec",
                "#2f7d32",
                "💸 할인 진행",
                f"시작: {latest_discount['discount_start_date']}<br>"
                f"종료: {latest_discount['discount_end_date']}"
            ))

        if not df_life.empty:
            new_events = df_life[df_life["lifecycle_event"] == "NEW_PRODUCT"]
            if not new_events.empty:
                latest_new = new_events.sort_values("date", ascending=False).iloc[0]
                cards.append(render_card(
                    bg="#f6f1e6",
                    border="#c88a00",
                    title="🆕 신제품",
                    content=f"발견일: {latest_new['date'].date()}"
                ))

        if not df_life.empty:
            restore_events = df_life[df_life["lifecycle_event"] == "RESTOCK"]
            if not restore_events.empty:
                restore_dates_str = "<br>".join([
                    f"날짜: {r['date'].date()}"
                    for _, r in restore_events.sort_values("date", ascending=False).iterrows()
                ])
                cards.append(render_card(
                    bg="#fff8e1",
                    border="#f59e0b",
                    title="🔄 복원",
                    content=restore_dates_str
                ))

        raw_res = (
            supabase.table("raw_daily_prices")
            .select("date, normal_price")
            .eq("product_url", p["product_url"])
            .gte("date", filter_date_from.strftime("%Y-%m-%d"))
            .lte("date", filter_date_to.strftime("%Y-%m-%d"))
            .order("date", desc=False)
            .execute()
        )
        if raw_res.data:
            raw_df = pd.DataFrame(raw_res.data)
            raw_df["normal_price"] = raw_df["normal_price"].astype(float)
            raw_df["date"] = pd.to_datetime(raw_df["date"])

            out_rows = raw_df[raw_df["normal_price"] == 0].copy()
            out_rows["prev_normal"] = raw_df["normal_price"].shift(1)
            out_start_rows = out_rows[out_rows["prev_normal"] != 0]
            lifecycle_out_dates = []
            if not df_life.empty:
                lifecycle_out_dates = [
                    str(d.date()) for d in df_life[df_life["lifecycle_event"] == "OUT_OF_STOCK"]["date"].tolist()
                ]
            missing_out = out_start_rows[~out_start_rows["date"].dt.strftime("%Y-%m-%d").isin(lifecycle_out_dates)]

            all_out_dates = lifecycle_out_dates + [
                str(r["date"].date()) for _, r in missing_out.iterrows()
            ]
            all_out_dates = sorted(list(set(all_out_dates)), reverse=True)
            all_out_dates = [d for d in all_out_dates if d and str(d) != "NaT"]

            if all_out_dates and not any("품절" in c for c in cards):
                out_dates_str = "<br>".join([f"날짜: {d}" for d in all_out_dates])
                cards.append(render_card(
                    bg="#e8f0f8",
                    border="#2c5aa0",
                    title="❌ 품절",
                    content=out_dates_str
                ))

            raw_df["prev_price"] = raw_df["normal_price"].shift(1)
            restore_rows = raw_df[(raw_df["prev_price"] == 0) & (raw_df["normal_price"] > 0)]
            if not restore_rows.empty and not any("복원" in c for c in cards):
                restore_dates_str = "<br>".join([f"날짜: {r['date'].date()}" for _, r in restore_rows.iterrows()])
                cards.append(render_card(
                    bg="#fff8e1",
                    border="#f59e0b",
                    title="🔄 복원",
                    content=restore_dates_str
                ))

        if normal_change_rows:
            latest_change = normal_change_rows[0]
            prev_price = float(latest_change["prev_price"])
            current_price = float(latest_change["normal_price"])

            if current_price == 0:
                already_has_out = any("품절" in c for c in cards)
                if not already_has_out:
                    cards.append(render_card(
                        bg="#e8f0f8",
                        border="#2c5aa0",
                        title="❌ 품절",
                        content=f"날짜: {latest_change['date']}<br>정상가 {prev_price:,.0f}원 → 품절"
                    ))
            elif prev_price == 0 and current_price > 0:
                already_has_restore = any("복원" in c for c in cards)
                if not already_has_restore:
                    cards.append(render_card(
                        bg="#fff8e1",
                        border="#f59e0b",
                        title="🔄 복원",
                        content=f"날짜: {latest_change['date']}<br>품절 → 정상가 {current_price:,.0f}원"
                    ))
            else:
                diff = current_price - prev_price
                diff_rate = (diff / prev_price) * 100 if prev_price != 0 else 0
                if diff > 0:
                    bg = "#fdecea"
                    border = "#b91c1c"
                    icon = "📈 정상가 상승"
                else:
                    bg = "#eaf2ff"
                    border = "#1d4ed8"
                    icon = "📉 정상가 하락"

                # ✅ capsule_count로 나눠서 개당 가격 표시
                cc = float(p.get("capsule_count") or 0)
                if cc > 0:
                    prev_unit = prev_price / cc
                    curr_unit = current_price / cc
                    price_text = (
                        f"{prev_unit:,.1f}원 → {curr_unit:,.1f}원 "
                        f"({diff_rate:+.1f}%)"
                    )
                else:
                    price_text = (
                        f"{prev_price:,.0f}원 → {current_price:,.0f}원 "
                        f"({diff_rate:+.1f}%)"
                    )

                cards.append(render_card(
                    bg=bg,
                    border=border,
                    title=icon,
                    content=(
                        f"날짜: {latest_change['date']}<br>"
                        f"{price_text}"
                    )
                ))
        if not cards:
            cards.append(render_card(
                "#f3f4f6",
                "#9aa0a6",
                "📊 특이 이벤트 없음",
                ""
            ))

        for row_start in range(0, len(cards), 3):
            row_cards = cards[row_start:row_start + 3]
            _, col1, col2, col3 = st.columns(4)
            for i, col in enumerate([col1, col2, col3]):
                if i < len(row_cards):
                    with col:
                        st.markdown(row_cards[i], unsafe_allow_html=True)

        st.markdown("<br><br>", unsafe_allow_html=True)

        with st.expander("📅 이벤트 히스토리"):

            display_rows = []

            discount_res = supabase.rpc(
                "get_discount_periods_in_range",
                {
                    "p_product_url": p["product_url"],
                    "p_date_from": filter_date_from.strftime("%Y-%m-%d"),
                    "p_date_to": filter_date_to.strftime("%Y-%m-%d"),
                }
            ).execute()

            discount_rows = discount_res.data if discount_res.data else []

            df_life = load_lifecycle_events(p["product_url"])
            out_dates = []
            if not df_life.empty:
                out_dates = pd.to_datetime(
                    df_life[df_life["lifecycle_event"] == "OUT_OF_STOCK"]["date"]
                ).tolist()

            for row in discount_rows:
                discount_start = pd.to_datetime(row["discount_start_date"])
                discount_end = pd.to_datetime(row["discount_end_date"])

                discount_price_res = (
                    supabase.table("product_all_events")
                    .select("unit_price")
                    .eq("product_url", p["product_url"])
                    .eq("event_type", "DISCOUNT")
                    .eq("date", row["discount_start_date"])
                    .limit(1)
                    .execute()
                )
                discount_price = (
                    float(discount_price_res.data[0]["unit_price"])
                    if discount_price_res.data else None
                )
                display_rows.append({
                    "날짜": row["discount_start_date"],
                    "이벤트": "💸 할인 시작",
                    "가격 정보": f"할인가 {discount_price:,.1f}원" if discount_price else ""
                })

                show_discount_end = True
                for out_date in out_dates:
                    if discount_end + pd.Timedelta(days=1) == out_date:
                        show_discount_end = False
                        break

                if show_discount_end:
                    discount_end_price_res = (
                        supabase.table("product_all_events")
                        .select("unit_price")
                        .eq("product_url", p["product_url"])
                        .eq("event_type", "DISCOUNT")
                        .eq("date", row["discount_end_date"])
                        .limit(1)
                        .execute()
                    )
                    discount_end_price = (
                        float(discount_end_price_res.data[0]["unit_price"])
                        if discount_end_price_res.data else None
                    )
                    display_rows.append({
                        "날짜": row["discount_end_date"],
                        "이벤트": "💸 할인 종료",
                        "가격 정보": f"종료가 {discount_end_price:,.1f}원" if discount_end_price else ""
                    })

            df_life_all = load_lifecycle_events(p["product_url"])
            if not df_life_all.empty:
                df_life_all["date"] = pd.to_datetime(df_life_all["date"], errors="coerce")
                df_life_all = df_life_all.dropna(subset=["date"])
                df_life = df_life_all.copy()
                df_life = df_life[
                    (df_life["date"] >= pd.Timestamp(filter_date_from)) &
                    (df_life["date"] <= pd.Timestamp(filter_date_to))
                ]
            else:
                df_life = pd.DataFrame(columns=["date", "lifecycle_event"])

            lifecycle_map = {
                "NEW_PRODUCT": "🆕 신제품",
                "OUT_OF_STOCK": "❌ 품절",
                "RESTOCK": "🔄 복원",
            }

            for _, row in df_life.iterrows():
                event_date = row["date"]
                event_type = row["lifecycle_event"]
                price_info = ""
                if event_type == "OUT_OF_STOCK":
                    price_res = (
                        supabase.table("product_all_events")
                        .select("unit_price, event_type")
                        .eq("product_url", p["product_url"])
                        .lt("date", event_date.strftime("%Y-%m-%d"))
                        .order("date", desc=True)
                        .limit(1)
                        .execute()
                    )
                    if price_res.data:
                        prev = float(price_res.data[0]["unit_price"])
                        etype = price_res.data[0]["event_type"]
                        price_label = "할인가" if etype == "DISCOUNT" else "정상가"
                        price_info = f"{price_label} {prev:,.1f}원 → 품절"
                elif event_type == "RESTOCK":
                    price_res = (
                        supabase.table("product_all_events")
                        .select("unit_price, event_type")
                        .eq("product_url", p["product_url"])
                        .gte("date", event_date.strftime("%Y-%m-%d"))
                        .order("date", desc=False)
                        .limit(1)
                        .execute()
                    )
                    if price_res.data:
                        after = float(price_res.data[0]["unit_price"])
                        etype = price_res.data[0]["event_type"]
                        price_label = "할인가" if etype == "DISCOUNT" else "정상가"
                        price_info = f"품절 → {price_label} {after:,.1f}원"

                display_rows.append({
                    "날짜": event_date.strftime("%Y-%m-%d"),
                    "이벤트": lifecycle_map.get(event_type, ""),
                    "가격 정보": price_info
                })

            df_life = load_lifecycle_events(p["product_url"])
            if not df_life.empty:
                df_life["date"] = pd.to_datetime(df_life["date"], errors="coerce")
                df_life = df_life.dropna(subset=["date"])
                df_life = df_life[
                    (df_life["date"] >= pd.Timestamp(filter_date_from)) &
                    (df_life["date"] <= pd.Timestamp(filter_date_to))
                ]

            df_changes_res = (
                supabase.table("product_price_change_events")
                .select("*")
                .eq("product_url", p["product_url"])
                .gte("date", filter_date_from.strftime("%Y-%m-%d"))
                .lte("date", filter_date_to.strftime("%Y-%m-%d"))
                .order("date", desc=True)
                .execute()
            )

            df_changes = pd.DataFrame(df_changes_res.data)

            if not df_changes.empty:
                icon_map = {
                    "DISCOUNT_DOWN": "💸 할인가 하락",
                    "DISCOUNT_UP": "💸 할인가 상승",
                    "NORMAL_UP": "📈 정상가 상승",
                    "NORMAL_DOWN": "📉 정상가 하락",
                }

                restore_dates_in_display = []
                if not df_life_all.empty:
                    restore_dates_in_display = [
                        str(d.date())
                        for d in df_life_all[df_life_all["lifecycle_event"] == "RESTOCK"]["date"].tolist()
                    ]
                discount_end_dates = [
                    r["날짜"] for r in display_rows if r["이벤트"] == "💸 할인 종료"
                ]
                restore_dates_in_display += discount_end_dates
                discount_end_dates_plus1 = [
                    str((pd.Timestamp(d) + pd.Timedelta(days=1)).date())
                    for d in discount_end_dates
                ]
                restore_dates_in_display += discount_end_dates_plus1

                for _, row in df_changes.iterrows():
                    prev_price = float(row["prev_price"]) if row["prev_price"] else 0
                    current_price = float(row["unit_price"]) if row["unit_price"] else 0

                    if current_price == 0 and row["price_change_type"] in ("NORMAL_DOWN", "NORMAL_UP"):
                        display_rows.append({
                            "날짜": row["date"],
                            "이벤트": "❌ 품절",
                            "가격 정보": f"정상가 {prev_price:,.1f}원 → 품절"
                        })
                        continue

                    if prev_price == 0 and current_price > 0 and row["price_change_type"] in ("NORMAL_DOWN", "NORMAL_UP"):
                        display_rows.append({
                            "날짜": row["date"],
                            "이벤트": "🔄 복원",
                            "가격 정보": f"품절 → 정상가 {current_price:,.1f}원"
                        })
                        continue

                    if row["price_change_type"] == "NORMAL_UP" and str(row["date"]) in restore_dates_in_display:
                        continue

                    if prev_price > 0:
                        diff = current_price - prev_price
                        diff_rate = (diff / prev_price) * 100
                        rate_text = f"({diff_rate:+.1f}%)"
                    else:
                        rate_text = ""

                    # ✅ NORMAL_UP / NORMAL_DOWN은 개당 가격으로 표시
                    price_text = f"{prev_price:,.1f}원 → {current_price:,.1f}원 {rate_text}"
                    _event = icon_map.get(row["price_change_type"], "")
                    if any(r["날짜"] == row["date"] and r["이벤트"] == _event for r in display_rows):
                        continue
                    display_rows.append({
                        "날짜": row["date"],
                        "이벤트": _event,
                        "가격 정보": price_text
                    })

            normal_res = (
                supabase.table("product_normal_price_events")
                .select("*")
                .eq("product_url", p["product_url"])
                .gte("date", filter_date_from.strftime("%Y-%m-%d"))
                .lte("date", filter_date_to.strftime("%Y-%m-%d"))
                .execute()
            )

            normal_rows = normal_res.data if normal_res.data else []

            for row in normal_rows:
                prev_price = float(row["prev_price"])
                current_price = float(row["normal_price"])

                if current_price == 0:
                    display_rows.append({
                        "날짜": row["date"],
                        "이벤트": "❌ 품절",
                        "가격 정보": f"정상가 {prev_price:,.1f}원 → 품절"
                    })
                    continue

                if prev_price == 0 and current_price > 0:
                    display_rows.append({
                        "날짜": row["date"],
                        "이벤트": "🔄 복원",
                        "가격 정보": f"품절 → 정상가 {current_price:,.1f}원"
                    })
                    continue

                diff = current_price - prev_price
                diff_rate = (diff / prev_price) * 100 if prev_price != 0 else 0
                event_label = "📈 정상가 상승" if diff > 0 else "📉 정상가 하락"

                # ✅ capsule_count로 나눠서 개당 가격 표시
                cc = float(p.get("capsule_count") or 0)
                if cc > 0:
                    prev_unit = prev_price / cc
                    curr_unit = current_price / cc
                    price_text = f"{prev_unit:,.1f}원 → {curr_unit:,.1f}원 ({diff_rate:+.1f}%)"
                else:
                    price_text = f"{prev_price:,.1f}원 → {current_price:,.1f}원 ({diff_rate:+.1f}%)"

                if any(r["날짜"] == row["date"] and r["이벤트"] == event_label for r in display_rows):
                    continue
                display_rows.append({
                    "날짜": row["date"],
                    "이벤트": event_label,
                    "가격 정보": price_text
                })
            if display_rows:
                df_display = pd.DataFrame(display_rows)
                df_display["날짜_정렬용"] = pd.to_datetime(df_display["날짜"], errors="coerce")
                # 조회 기간 필터 적용
                df_display = df_display[
                    (df_display["날짜_정렬용"] >= pd.Timestamp(filter_date_from)) &
                    (df_display["날짜_정렬용"] <= pd.Timestamp(filter_date_to))
                ]
                df_display = df_display.sort_values("날짜_정렬용", ascending=False)
                df_display = df_display.drop(columns=["날짜_정렬용"])
                st.dataframe(df_display, use_container_width=True, hide_index=True)
            else:
                st.caption("이벤트 없음")








